import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from io import BytesIO

st.set_page_config(page_title="Dcyphr Sale & Stock Report", layout="wide", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Plus+Jakarta+Sans:wght@600;700;800&display=swap');
*, *::before, *::after { font-family: 'Inter', sans-serif !important; box-sizing: border-box; }
.stApp { background: #f4f0ff !important; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 0.8rem !important; padding-bottom: 1rem !important; }
.hero {
    padding: 0.55rem 1.4rem; display: flex; align-items: center; gap: 1rem;
    background: linear-gradient(90deg, #3a0068 0%, #6a1b9a 55%, #9c27b0 100%);
    margin-bottom: 1rem; border-radius: 12px;
    box-shadow: 0 3px 14px rgba(106,27,154,0.3);
}
.hero-badge {
    background: rgba(255,255,255,0.18); border: 1.5px solid rgba(255,255,255,0.35);
    color: #ffffff; font-size:.56rem; font-weight:700; letter-spacing:2px;
    text-transform:uppercase; padding:4px 11px; border-radius:20px; white-space:nowrap;
}
.hero-title { font-family:'Plus Jakarta Sans',sans-serif !important; font-size:1.05rem; font-weight:800; color:#ffffff; }
.hero-sub-line { font-family:'Plus Jakarta Sans',sans-serif !important; font-size:.8rem; font-weight:600; color:#e8c8ff; }
.hero-sub { color:rgba(255,255,255,0.52); font-size:.65rem; font-weight:400; }
.kpi-card {
    background: linear-gradient(135deg, #6a1b9a 0%, #9c27b0 100%);
    border-radius: 16px; padding: 1.1rem 1.3rem;
    box-shadow: 0 4px 18px rgba(106,27,154,0.35);
}
.kpi-label { font-size:.58rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase; color:rgba(255,255,255,0.75); margin-bottom:.4rem; }
.kpi-value { font-family:'Plus Jakarta Sans',sans-serif !important; font-size:1.3rem; font-weight:800; color:#ffffff; line-height:1.1; }
.kpi-sub { font-size:.72rem; color:rgba(255,255,255,0.7); margin-top:.25rem; }
.section-title {
    font-size:.63rem; font-weight:700; letter-spacing:2.5px; text-transform:uppercase;
    color:#6a1b9a; padding:.4rem 0; margin-bottom:.7rem; border-bottom: 2px solid #ddd6fe;
}
p { color:#1a0030 !important; font-size:.9rem !important; }
label { color:#3d0066 !important; font-weight:600 !important; }
[data-testid="stWidgetLabel"] p { color:#6a1b9a !important; font-weight:700 !important; font-size:.85rem !important; }
.stSelectbox > div > div { background:#ffffff !important; border:1.5px solid #c084fc !important; border-radius:10px !important; color:#1a0030 !important; }
[data-baseweb="select"] > div { background:#ffffff !important; border:1.5px solid #c084fc !important; border-radius:10px !important; }
[data-baseweb="popover"] { background:#fff !important; border:1px solid #c084fc !important; }
[data-baseweb="popover"] * { color:#1a0030 !important; background:#fff !important; }
.stMultiSelect > div > div { background:#ffffff !important; border:1.5px solid #c084fc !important; border-radius:10px !important; }
.stButton > button {
    background: linear-gradient(135deg,#6a1b9a,#9c27b0) !important;
    color:#ffffff !important; border:none !important; border-radius:12px !important;
    font-weight:700 !important; padding:.7rem 2rem !important;
    box-shadow:0 4px 14px rgba(106,27,154,0.38) !important;
}
.stButton > button p { color:#ffffff !important; font-weight:700 !important; }
.stDownloadButton > button {
    background:#fff !important; color:#6a1b9a !important;
    border:2px solid #6a1b9a !important; border-radius:10px !important; font-weight:700 !important;
}
.stTabs [data-baseweb="tab-list"] {
    background:#fff !important; border-radius:12px !important; padding:4px !important;
    border:1.5px solid #ddd6fe !important;
}
.stTabs [data-baseweb="tab"] { color:#1a0030 !important; border-radius:8px !important; font-size:.82rem !important; font-weight:600 !important; }
.stTabs [aria-selected="true"] { background:linear-gradient(135deg,#6a1b9a,#9c27b0) !important; color:#ffffff !important; }
.stTabs [aria-selected="true"] * { color:#ffffff !important; }
div[data-testid="stDataFrame"] * { color:#1a0030 !important; font-size:.84rem !important; }
[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important; border: 2px dashed #c084fc !important;
    border-radius: 14px !important; padding: 16px !important;
    display: flex !important; flex-direction: column !important;
    align-items: center !important; justify-content: center !important;
    min-height: 0 !important; max-height: 80px !important;
}
[data-testid="stFileUploaderDropzone"] button {
    visibility: hidden !important; height: 0 !important; padding: 0 !important; margin: 0 !important;
}
</style>
""", unsafe_allow_html=True)

import re
MONTH_MAP = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'june':6,
    'jul':7,'july':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12,
    'march':3,'april':4,'august':8,'september':9,'october':10,
    'november':11,'december':12,'january':1,'february':2
}
MONTH_SHORT_MAP = {
    'jan':'Jan','feb':'Feb','mar':'Mar','apr':'Apr','may':'May','jun':'Jun',
    'june':'Jun','jul':'Jul','july':'Jul','aug':'Aug','sep':'Sep',
    'oct':'Oct','nov':'Nov','dec':'Dec','march':'Mar','april':'Apr',
    'august':'Aug','september':'Sep','october':'Oct','november':'Nov',
    'december':'Dec','january':'Jan','february':'Feb'
}

def parse_month_order(months_series):
    parsed = []
    seen = set()
    for m in months_series.dropna().unique():
        m = str(m).strip()
        if m in seen: continue
        seen.add(m)
        match = re.match(r"([A-Za-z]+)['\s]?(\d{2,4})?", m)
        if match:
            mon_str = match.group(1).lower()
            yr_str  = match.group(2) or '25'
            yr = int(yr_str) if len(yr_str)==4 else 2000+int(yr_str)
            mon_num = MONTH_MAP.get(mon_str, 0)
            if mon_num > 0:
                parsed.append((yr*100 + mon_num, m))
    parsed.sort(key=lambda x: x[0])
    return [x[1] for x in parsed]

def get_month_short(m):
    match = re.match(r"([A-Za-z]+)['\s]?(\d{2,4})?", str(m).strip())
    if match:
        base = MONTH_SHORT_MAP.get(match.group(1).lower(), match.group(1)[:3].capitalize())
        yr = match.group(2)
        if yr and len(yr)==2: return base + "'" + yr
        return base
    return str(m)[:3]

CAT_COLORS = ['#7b1fa2','#e91e63','#ff6f00','#1565c0','#2e7d32','#00838f','#f57f17','#6a1b9a',
              '#c62828','#00695c','#4527a0','#ad1457','#558b2f','#0277bd']
BLUE_SEQ   = [[0,'#f3e5f5'],[0.4,'#9c27b0'],[1,'#6a1b9a']]
SIZE_ORDER = ['XS','S','M','L','XL','XXL','XXXL','2XL','3XL']

def fmt_inr(v):
    if pd.isna(v) or v == 0: return "—"
    v = int(round(float(v)))
    s = str(abs(v)); prefix = "-" if v < 0 else ""
    if len(s) <= 3: return prefix + s
    last3 = s[-3:]; rest = s[:-3]; groups = []
    while len(rest) > 2: groups.append(rest[-2:]); rest = rest[:-2]
    if rest: groups.append(rest)
    groups.reverse()
    return prefix + ','.join(groups) + ',' + last3

def pct(v, dec=1):
    if pd.isna(v) or v == 0: return "—"
    return f"{float(v)*100:.{dec}f}%"

def cl(height=380, title="", xangle=0, margin=None):
    m = margin or dict(l=10, r=10, t=55, b=40)
    return dict(
        paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
        font=dict(color="#1a0030", family="Inter", size=12),
        margin=m, height=height,
        title=dict(text=f"<b>{title}</b>", font=dict(color="#1a0030", size=14, family="Plus Jakarta Sans")),
        legend=dict(font=dict(color="#1a0030", size=11), bgcolor="rgba(255,255,255,0.97)",
                    bordercolor="#ddd6fe", borderwidth=1.5),
        xaxis=dict(gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11),
                   linecolor="#ddd6fe", tickangle=xangle, showgrid=True),
        yaxis=dict(gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11),
                   linecolor="#ddd6fe", showgrid=True),
    )

def kpi(col, label, value, sub, icon):
    with col:
        st.markdown(f"""<div class="kpi-card">
          <div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:.3rem">
            <div class="kpi-label">{label}</div><span style="font-size:1.1rem">{icon}</span>
          </div>
          <div class="kpi-value">{value}</div>
          <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

def sec(title):
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def process(file_bytes):
    df = pd.read_excel(file_bytes, sheet_name='Sale data ', header=0)
    df.columns = [str(c).strip() for c in df.columns]
    df = df.rename(columns={
        'DEVISION DESC.': 'Division',
        'CATEGORY DESC.': 'Category',
        'GENDER.': 'Gender',
        'Brand.': 'Brand',
        'Net Sale (Tax Incl)': 'NetSale',
        'Month Name': 'Month',
        'Sub Category Code': 'SubCategory',
        'SUB CATEGORY DESC.': 'SubCategoryDesc',
        'Sub Category2 Code.': 'FitType',
        'Season sub Group': 'Season',
        'Season Desc.': 'SeasonDesc',
    })
    for col in ['Sale Qty', 'NetSale', 'MRP Value', 'Discount Amount', 'MRP']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    df['Distributor'] = df['Distributor'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
    df['Store Name']  = df['Store Name'].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
    df['Gender']      = df['Gender'].astype(str).str.upper().str.strip()
    df['Category']    = df['Category'].astype(str).str.upper().str.strip()
    df['Color']       = df['Color'].astype(str).str.upper().str.strip() if 'Color' in df.columns else 'N/A'
    df['Size']        = df['Size'].astype(str).str.upper().str.strip() if 'Size' in df.columns else 'N/A'
    df['Item Name']   = df['Item Name'].astype(str).str.strip() if 'Item Name' in df.columns else df['Item ID']
    df['SubCategory'] = df['SubCategory'].astype(str).str.strip() if 'SubCategory' in df.columns else 'N/A'
    df['SubCategoryDesc'] = df['SubCategoryDesc'].astype(str).str.strip() if 'SubCategoryDesc' in df.columns else 'N/A'
    df['FitType']     = df['FitType'].astype(str).str.strip() if 'FitType' in df.columns else 'N/A'
    df['Season']      = df['Season'].astype(str).str.strip() if 'Season' in df.columns else 'N/A'
    months_order = parse_month_order(df['Month'])
    month_short  = [get_month_short(m) for m in months_order]
    return df, months_order, month_short


@st.cache_data(show_spinner=False)
def process_store(file_bytes):
    sale  = pd.read_excel(file_bytes, sheet_name='BILL WISE SALE REPORT APRIL2025', header=0)
    stock = pd.read_excel(file_bytes, sheet_name='DCYPHR STORE STOCK REPORT', header=0)
    sale.columns  = [str(c).strip() for c in sale.columns]
    stock.columns = [str(c).strip() for c in stock.columns]
    sale = sale.rename(columns={
        'Store Name': 'StoreName', 'STORE CODE': 'StoreCode',
        'Item No/Article Code': 'ItemID', 'Item Name': 'ItemName',
        'Divison  Desc': 'DivisionDesc', 'Category Desc': 'Category',
        'Sub Category Code': 'SubCategory', 'Sub Category2 Code.': 'FitType',
        'GENDER.': 'Gender', 'Season sub Group': 'Season', 'Brand.': 'Brand',
        'NET_SALE_TAX _INCL': 'NetSale', 'MRP VALUE': 'MRPValue',
        'DISCOUNT AMOUNT': 'DiscAmt', 'TOTAL QTY': 'SaleQty',
        'Column1': 'Month', 'Color': 'Color', 'Size': 'Size',
    })
    stock = stock.rename(columns={
        'Store Name': 'StoreName', 'STORE CODE': 'StoreCode',
        'Item ID': 'ItemID', 'Item Name': 'ItemName',
        'Divison  Desc': 'Division', 'Category Desc': 'Category',
        'Sub Category Code': 'SubCategory', 'Sub Category2 Code.': 'FitType',
        'GENDER.': 'Gender', 'Season sub Group': 'Season', 'Brand.': 'Brand',
        'MRP Value': 'MRPValue', 'Discount Amount': 'DiscAmt',
        'Closing Qty': 'ClosingQty', 'Clsoing Value': 'ClosingValue',
        'Color': 'Color', 'Size': 'Size',
    })
    for col in ['NetSale','MRPValue','DiscAmt','SaleQty']:
        if col in sale.columns:
            sale[col] = pd.to_numeric(sale[col], errors='coerce').fillna(0)
    for col in ['MRPValue','DiscAmt','ClosingQty','ClosingValue']:
        if col in stock.columns:
            stock[col] = pd.to_numeric(stock[col], errors='coerce').fillna(0)
    sale['StoreName'] = sale['StoreName'].astype(str).str.strip()
    sale['Gender']    = sale['Gender'].astype(str).str.upper().str.strip()
    sale['Category']  = sale['Category'].astype(str).str.upper().str.strip()
    sale['Season']    = sale['Season'].astype(str).str.strip()
    sale['Color']     = sale['Color'].astype(str).str.upper().str.strip()
    sale['Size']      = sale['Size'].astype(str).str.upper().str.strip()
    sale['Month']     = sale['Month'].astype(str).str.strip()
    sale['ItemName']  = sale['ItemName'].astype(str).str.strip() if 'ItemName' in sale.columns else sale['ItemID']
    sale['SubCategory'] = sale['SubCategory'].astype(str).str.strip() if 'SubCategory' in sale.columns else 'N/A'
    sale['FitType']   = sale['FitType'].astype(str).str.strip() if 'FitType' in sale.columns else 'N/A'
    stock['StoreName'] = stock['StoreName'].astype(str).str.strip()
    stock['Gender']    = stock['Gender'].astype(str).str.upper().str.strip()
    stock['Category']  = stock['Category'].astype(str).str.upper().str.strip()
    stock['Season']    = stock['Season'].astype(str).str.strip()
    stock['Color']     = stock['Color'].astype(str).str.upper().str.strip()
    stock['Size']      = stock['Size'].astype(str).str.upper().str.strip()
    STORE_MONTH_ORDER = ['Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec','Jan','Feb','Mar']
    present_months = [m for m in STORE_MONTH_ORDER if m in sale['Month'].unique()]
    xl_tmp = pd.ExcelFile(file_bytes)
    wh_sheet = 'Sheet2'
    for sname in xl_tmp.sheet_names:
        if 'warehouse' in sname.lower() or 'inventory' in sname.lower() or sname == 'Sheet2':
            wh_sheet = sname
            break
    file_bytes.seek(0)
    wh = pd.read_excel(file_bytes, sheet_name=wh_sheet, header=0)
    wh.columns = [str(c).strip() for c in wh.columns]
    if 'Season' in wh.columns and 'Season Sub Group' in wh.columns:
        wh = wh.rename(columns={'Season': 'SeasonDesc'})
    wh = wh.rename(columns={
        'LOCATION NAME': 'Location', 'STORE CODE': 'StoreCode',
        'ITEM NO/ARTICLE CODE': 'ItemID', 'ITEM NAME': 'ItemName',
        'COLOR': 'Color', 'SIZE': 'Size',
        'Divison  Desc': 'Division', 'Category Desc': 'Category',
        'Sub Category Desc': 'SubCategory', 'Sub Category2 Code.': 'FitType',
        'Gender': 'Gender', 'Season Sub Group': 'Season', 'Brand': 'Brand',
        'MRP/UNIT': 'MRP', 'CLOSING STOCK QTY': 'ClosingQty',
        'CLOSING VALUE(MRP)': 'ClosingValueMRP',
        'CLOSING VALUE( COST PRICE WITHOUT TAX)': 'ClosingValueCost',
        'GODOWN NAME': 'GodownName',
    })
    for col in ['MRP','ClosingQty','ClosingValueMRP','ClosingValueCost']:
        if col in wh.columns:
            wh[col] = pd.to_numeric(wh[col], errors='coerce').fillna(0)
    wh['Color']    = wh['Color'].astype(str).str.upper().str.strip() if 'Color' in wh.columns else 'N/A'
    wh['Size']     = wh['Size'].astype(str).str.upper().str.strip() if 'Size' in wh.columns else 'N/A'
    wh['Gender']   = wh['Gender'].astype(str).str.upper().str.strip() if 'Gender' in wh.columns else 'N/A'
    wh['Season']   = wh['Season'].astype(str).str.strip() if 'Season' in wh.columns else 'N/A'
    wh['Category'] = wh['Category'].astype(str).str.upper().str.strip() if 'Category' in wh.columns else 'N/A'
    wh['ItemName'] = wh['ItemName'].astype(str).str.strip() if 'ItemName' in wh.columns else 'N/A'
    wh['GodownName'] = wh['GodownName'].astype(str).str.strip() if 'GodownName' in wh.columns else 'N/A'
    wh = wh.dropna(subset=['GodownName'])
    wh = wh[wh['GodownName'].str.strip() != 'nan']
    return sale, stock, wh, present_months

for k,v in {"ready":False,"data":None,"mode":None,"store_data":None}.items():
    if k not in st.session_state: st.session_state[k] = v

st.markdown("""
<div class="hero">
  <div class="hero-badge">DCYPHR</div>
  <div style="width:1px;height:26px;background:rgba(255,255,255,.22)"></div>
  <div>
    <div style="display:flex;align-items:baseline;gap:.6rem">
      <div class="hero-title">Dcyphr Sale &amp; Stock Report</div>
      <div style="color:rgba(255,255,255,.4)">→</div>
      <div class="hero-sub-line">Channel · Store · Article · Colour · Size · Season Analysis</div>
    </div>
    <div class="hero-sub">Upload Report &nbsp;·&nbsp; Auto Analysis &nbsp;·&nbsp; Excel Export</div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""<div style="text-align:center;margin-bottom:1rem">
    <div style="font-size:.75rem;font-weight:600;color:#6a1b9a;margin-bottom:.6rem;text-transform:uppercase;letter-spacing:1px">
        Select Report Type to Upload
    </div>
</div>""", unsafe_allow_html=True)

b1, b2 = st.columns(2)
with b1:
    if st.button("📦  Channel Sale & Stock Report", use_container_width=True):
        st.session_state.mode = "channel"
        st.session_state.ready = False
        st.session_state.data = None
with b2:
    if st.button("🏪  Store Sale & Stock Report", use_container_width=True):
        st.session_state.mode = "store"
        st.session_state.ready = False
        st.session_state.data = None
        st.session_state.store_data = None

if st.session_state.mode:
    mode_label = "Channel Sale & Stock Report" if st.session_state.mode == "channel" else "Store Sale & Stock Report"
    u1,u2,u3 = st.columns([1,2,1])
    with u2:
        uploaded = st.file_uploader(
            f"📂 Upload {mode_label} · XLSX",
            type=["xlsx","xls"],
            label_visibility="visible",
            key=f"up_{st.session_state.mode}"
        )
        if uploaded:
            if st.button("⚡  Generate Analysis", use_container_width=True):
                with st.spinner("Processing..."):
                    file_bytes = BytesIO(uploaded.read())
                    if st.session_state.mode == "channel":
                        st.session_state.data  = process(file_bytes)
                    else:
                        st.session_state.store_data = process_store(file_bytes)
                    st.session_state.ready = True
                st.success("✅ Done! Analysis Ready.")

if not st.session_state.ready:
    st.markdown("""<div style="text-align:center;padding:4rem 0">
      <div style="font-size:3rem">📊</div>
      <div style="margin-top:1rem;font-size:1rem;color:#607d9b;font-weight:500">
        Select report type above and upload file</div>
      <div style="margin-top:.4rem;font-size:.82rem;color:#90a4c0">
        Channel wise or Store wise analysis will be generated instantly</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

mode = st.session_state.mode

if mode == "store" and st.session_state.store_data:
    sale_s, stock_s, wh_s, MONTHS_S = st.session_state.store_data
    SIZE_ORDER_S = ['XS','S','M','L','XL','XXL','XXXL','2XL','3XL','STANDARD']

    total_sale_s  = sale_s['NetSale'].sum()
    total_qty_s   = int(sale_s['SaleQty'].sum())
    total_stock_s = stock_s['ClosingValue'].sum()
    total_stk_qty_s = int(stock_s['ClosingQty'].sum())
    top_store_s   = sale_s.groupby('StoreName')['NetSale'].sum().idxmax() if len(sale_s)>0 else "—"

    k1,k2,k3,k4,k5 = st.columns(5)
    kpi(k1,"Total Net Sale",   f"₹{fmt_inr(int(total_sale_s))}", f"{MONTHS_S[0]}–{MONTHS_S[-1]}", "💰")
    kpi(k2,"Total Qty Sold",   f"{total_qty_s:,} Pcs",           "All stores", "📦")
    kpi(k3,"Closing Stock",    f"₹{fmt_inr(int(total_stock_s))}", f"{total_stk_qty_s:,} Pcs", "🏬")
    kpi(k4,"Total Stores",     str(sale_s['StoreName'].nunique()), "Active stores", "🏪")
    kpi(k5,"Top Store",        top_store_s[:20], f"₹{fmt_inr(int(sale_s.groupby('StoreName')['NetSale'].sum().max()))}", "🏆")
    st.markdown("<br>", unsafe_allow_html=True)

    st1,st2,st3,st4,st5,st6,st7,st8,st9,st10,st11 = st.tabs([
        "📈 Overview","🏪 Store-wise","🏷️ Article-wise",
        "🎨 Colour-wise","📐 Size-wise","👤 Gender+Category",
        "🗓️ Season-wise","📊 Sale vs Stock","🏭 Warehouse Stock","📋 Export","🎯 Final Review"
    ])

    with st1:
        sec("📈 Monthly Sale Trend")
        monthly_s = sale_s.groupby('Month')['NetSale'].sum().reindex(MONTHS_S).fillna(0)
        bi_s = int(monthly_s.values.argmax()); wi_s = int(monthly_s.values.argmin())
        bcolors_s = ['#9c27b0' if i not in [bi_s,wi_s] else ('#16a34a' if i==bi_s else '#dc2626') for i in range(len(monthly_s))]
        fig_s = go.Figure(go.Bar(x=MONTHS_S, y=monthly_s.values,
            marker=dict(color=bcolors_s, line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in monthly_s.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_s.update_layout(**cl(360,"Monthly Net Sale — All Stores",margin=dict(l=10,r=10,t=55,b=40)),
            bargap=0.3, yaxis_range=[0,monthly_s.max()*1.25])
        st.plotly_chart(fig_s, use_container_width=True)
        ca_s,cb_s = st.columns(2)
        with ca_s:
            sec("🏆 Top 10 Stores by Sale")
            top10_s = sale_s.groupby('StoreName')['NetSale'].sum().nlargest(10).sort_values()
            fig_t10 = go.Figure(go.Bar(x=top10_s.values, y=[s[:28] for s in top10_s.index], orientation='h',
                marker=dict(color=top10_s.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in top10_s.values],
                textposition='outside', textfont=dict(size=10,color='#1a0030')))
            fig_t10.update_layout(**cl(420,"Top 10 Stores",margin=dict(l=10,r=160,t=55,b=40)),
                xaxis_range=[0,top10_s.max()*1.45])
            st.plotly_chart(fig_t10, use_container_width=True)
        with cb_s:
            sec("🗂️ Category Mix")
            cat_s2 = sale_s.groupby('Category')['NetSale'].sum().sort_values(ascending=False)
            cat_s2 = cat_s2[cat_s2>0]
            fig_cs2 = go.Figure(go.Pie(labels=cat_s2.index.tolist(), values=cat_s2.values.tolist(), hole=0.5,
                marker=dict(colors=CAT_COLORS[:len(cat_s2)], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=11,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_cs2.update_layout(**cl(420,"Category-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
            st.plotly_chart(fig_cs2, use_container_width=True)

    with st2:
        sec("🏪 Store-wise Monthly Sale")
        stores_list = sorted(sale_s['StoreName'].dropna().unique().tolist())
        top5_s = sale_s.groupby('StoreName')['NetSale'].sum().nlargest(5).index.tolist()
        sel_stores_s = st.multiselect("Select Stores", stores_list, default=top5_s[:3], key="sw_sel_s")
        if sel_stores_s:
            fig_sw_s = go.Figure()
            for i,s in enumerate(sel_stores_s):
                sd = sale_s[sale_s['StoreName']==s].groupby('Month')['NetSale'].sum().reindex(MONTHS_S).fillna(0)
                fig_sw_s.add_trace(go.Bar(x=MONTHS_S, y=sd.values, name=s[:25],
                    marker_color=CAT_COLORS[i%len(CAT_COLORS)]))
            fig_sw_s.update_layout(**cl(400,"Store-wise Monthly Sale",margin=dict(l=10,r=10,t=55,b=40)),
                barmode='group', bargap=0.12)
            st.plotly_chart(fig_sw_s, use_container_width=True)
        sec("📋 Store Summary Table")
        st_tbl_s = sale_s.groupby('StoreName').agg(
            NetSale=('NetSale','sum'), Qty=('SaleQty','sum'),
            MRP=('MRPValue','sum'), Disc=('DiscAmt','sum')
        ).reset_index().sort_values('NetSale',ascending=False)
        st_tbl_s['Avg Disc%'] = (st_tbl_s['Disc']/st_tbl_s['MRP']*100).round(1)
        st_tbl_s['Sale Cont%'] = (st_tbl_s['NetSale']/st_tbl_s['NetSale'].sum()*100).round(2)
        st_tbl_s['NetSale'] = st_tbl_s['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
        st_tbl_s['MRP'] = st_tbl_s['MRP'].apply(lambda x: f"₹{fmt_inr(int(x))}")
        st_tbl_s['Qty'] = st_tbl_s['Qty'].apply(int)
        st_tbl_s = st_tbl_s.drop(columns=['Disc'])
        st_tbl_s.columns = ['Store','Net Sale','Qty','MRP Value','Avg Disc%','Sale Cont%']
        st.dataframe(st_tbl_s, use_container_width=True, hide_index=True)
        sec("📅 Store Month-wise Pivot")
        sw_piv = sale_s.pivot_table(index='StoreName',columns='Month',values='NetSale',aggfunc='sum').reindex(columns=MONTHS_S).fillna(0)
        sw_piv['Total'] = sw_piv.sum(axis=1)
        sw_piv = sw_piv.sort_values('Total',ascending=False)
        sw_piv = sw_piv.map(lambda x: f"₹{fmt_inr(int(x))}" if x!=0 else "—")
        st.dataframe(sw_piv, use_container_width=True)

    with st3:
        sec("🏷️ Article-wise Sale Analysis")
        af1_s,af2_s,af3_s,af4_s = st.columns(4)
        with af1_s: a_st_s = st.selectbox("Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="art_st_s")
        with af2_s: a_cat_s = st.multiselect("Category", sorted(sale_s['Category'].dropna().unique()), key="art_cat_s")
        with af3_s:
            cat_divs_s = sorted(sale_s[sale_s['Category'].isin(a_cat_s)]['DivisionDesc'].dropna().unique()) if a_cat_s else sorted(sale_s['DivisionDesc'].dropna().unique())
            a_div_s = st.multiselect("Division", cat_divs_s, key="art_div_s")
        with af4_s: a_srch_s = st.text_input("🔍 Search Item", placeholder="Item name...", key="art_srch_s")
        adf_s = sale_s.copy()
        if a_st_s != "All": adf_s = adf_s[adf_s['StoreName']==a_st_s]
        if a_cat_s: adf_s = adf_s[adf_s['Category'].isin(a_cat_s)]
        if a_div_s: adf_s = adf_s[adf_s['DivisionDesc'].isin(a_div_s)]
        if a_srch_s: adf_s = adf_s[adf_s['ItemName'].str.contains(a_srch_s, case=False, na=False)]
        top_arts_s = adf_s.groupby(['ItemID','ItemName'])['NetSale'].sum().sort_values(ascending=False).head(10)
        top_arts_sorted_s = top_arts_s.sort_values(ascending=True)
        art_title_parts = []
        if a_cat_s: art_title_parts.append(" + ".join(a_cat_s))
        if a_div_s: art_title_parts.append(" + ".join(a_div_s))
        art_chart_title = "Top 10 Articles by Net Sale"
        if art_title_parts: art_chart_title += f" — {' > '.join(art_title_parts)}"
        sec("🏆 Top 10 Articles by Net Sale")
        if len(top_arts_sorted_s)>0:
            y_labels = [f"{idx[1][:35]} [{idx[0]}]" for idx in top_arts_sorted_s.index]
            fig_art_s = go.Figure(go.Bar(
                x=top_arts_sorted_s.values, y=y_labels, orientation='h',
                marker=dict(color=top_arts_sorted_s.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in top_arts_sorted_s.values],
                textposition='outside', textfont=dict(size=11,color='#1a0030')))
            fig_art_s.update_layout(**cl(480,art_chart_title,margin=dict(l=10,r=160,t=55,b=40)),
                xaxis_range=[0,top_arts_sorted_s.max()*1.55])
            st.plotly_chart(fig_art_s, use_container_width=True)
        sec("📋 Article Summary Table")
        art_tbl_s = adf_s.groupby(['ItemID','ItemName','Category','Gender']).agg(
            NetSale=('NetSale','sum'), Qty=('SaleQty','sum'), Stores=('StoreName','nunique')
        ).reset_index().sort_values('NetSale',ascending=False)
        art_tbl_s['NetSale'] = art_tbl_s['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
        art_tbl_s['Qty'] = art_tbl_s['Qty'].apply(int)
        art_tbl_s.columns = ['Item ID','Item Name','Category','Gender','Net Sale','Qty','Stores']
        st.dataframe(art_tbl_s, use_container_width=True, hide_index=True)

    with st4:
        sec("🎨 Colour-wise Sale Analysis")
        cf1_s,cf2_s,cf3_s = st.columns(3)
        with cf1_s: c_st_s = st.selectbox("Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="col_st_s")
        with cf2_s: c_cat_s = st.multiselect("Category", sorted(sale_s['Category'].dropna().unique()), key="col_cat_s")
        with cf3_s:
            ccat_divs_s = sorted(sale_s[sale_s['Category'].isin(c_cat_s)]['DivisionDesc'].dropna().unique()) if c_cat_s else sorted(sale_s['DivisionDesc'].dropna().unique())
            c_div_s = st.multiselect("Division", ccat_divs_s, key="col_div_s")
        cdf_s = sale_s.copy()
        if c_st_s != "All": cdf_s = cdf_s[cdf_s['StoreName']==c_st_s]
        if c_cat_s: cdf_s = cdf_s[cdf_s['Category'].isin(c_cat_s)]
        if c_div_s: cdf_s = cdf_s[cdf_s['DivisionDesc'].isin(c_div_s)]
        col_sale_s = cdf_s.groupby('Color')['NetSale'].sum().sort_values(ascending=False).head(15)
        fig_col_s = go.Figure(go.Bar(x=col_sale_s.index.tolist(), y=col_sale_s.values,
            marker=dict(color=CAT_COLORS[:len(col_sale_s)], line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in col_sale_s.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_col_s.update_layout(**cl(320,"Top Colours by Net Sale",margin=dict(l=10,r=10,t=55,b=70)),
            bargap=0.3, xaxis_tickangle=-30, yaxis_range=[0,col_sale_s.max()*1.25])
        st.plotly_chart(fig_col_s, use_container_width=True)
        col_tbl_s = cdf_s.groupby(['Color','Category']).agg(
            NetSale=('NetSale','sum'), Qty=('SaleQty','sum')
        ).reset_index().sort_values('NetSale',ascending=False)
        col_tbl_s['NetSale'] = col_tbl_s['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
        col_tbl_s.columns = ['Colour','Category','Net Sale','Qty']
        st.dataframe(col_tbl_s, use_container_width=True, hide_index=True)

    with st5:
        sec("📐 Size-wise Sale Analysis")
        sf1_s,sf2_s,sf3_s,sf4_s = st.columns(4)
        with sf1_s: s_st_s = st.selectbox("Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="sz_st_s")
        with sf2_s: s_gen_s = st.selectbox("Gender", ["All"]+sorted(sale_s['Gender'].dropna().unique()), key="sz_gen_s")
        with sf3_s: s_cat_s = st.multiselect("Category", sorted(sale_s['Category'].dropna().unique()), key="sz_cat_s")
        with sf4_s:
            scat_divs_s = sorted(sale_s[sale_s['Category'].isin(s_cat_s)]['DivisionDesc'].dropna().unique()) if s_cat_s else sorted(sale_s['DivisionDesc'].dropna().unique())
            s_div_s = st.multiselect("Division", scat_divs_s, key="sz_div_s")
        sdf_s = sale_s.copy()
        if s_st_s != "All": sdf_s = sdf_s[sdf_s['StoreName']==s_st_s]
        if s_gen_s != "All": sdf_s = sdf_s[sdf_s['Gender']==s_gen_s]
        if s_cat_s: sdf_s = sdf_s[sdf_s['Category'].isin(s_cat_s)]
        if s_div_s: sdf_s = sdf_s[sdf_s['DivisionDesc'].isin(s_div_s)]
        all_sz_s = sdf_s['Size'].dropna().unique().tolist()
        ord_sz_s = [s for s in SIZE_ORDER_S if s in all_sz_s] + [s for s in all_sz_s if s not in SIZE_ORDER_S]
        sz_title_parts = []
        if s_cat_s: sz_title_parts.append(" + ".join(s_cat_s))
        if s_div_s: sz_title_parts.append(" + ".join(s_div_s))
        sz_filter_label = " > ".join(sz_title_parts) if sz_title_parts else "All"
        sa_s,sb_s = st.columns(2)
        active_divs_s = sdf_s['DivisionDesc'].dropna().unique().tolist()
        if len(active_divs_s) > 1:
            with sa_s:
                fig_szs_s = go.Figure()
                for i,div in enumerate(sorted(active_divs_s)):
                    ddf = sdf_s[sdf_s['DivisionDesc']==div]
                    dv = ddf.groupby('Size')['NetSale'].sum().reindex(ord_sz_s).fillna(0)
                    fig_szs_s.add_trace(go.Bar(name=div, x=ord_sz_s, y=dv.values,
                        marker_color=CAT_COLORS[i%len(CAT_COLORS)],
                        text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in dv.values],
                        textposition='outside', textfont=dict(size=9,color='#1a0030')))
                fig_szs_s.update_layout(**cl(360,f"Size-wise Net Sale — {sz_filter_label}",margin=dict(l=10,r=10,t=55,b=40)),
                    barmode='group', bargap=0.15)
                st.plotly_chart(fig_szs_s, use_container_width=True)
            with sb_s:
                fig_szq_s = go.Figure()
                for i,div in enumerate(sorted(active_divs_s)):
                    ddf = sdf_s[sdf_s['DivisionDesc']==div]
                    dq = ddf.groupby('Size')['SaleQty'].sum().reindex(ord_sz_s).fillna(0)
                    fig_szq_s.add_trace(go.Bar(name=div, x=ord_sz_s, y=dq.values,
                        marker_color=CAT_COLORS[i%len(CAT_COLORS)],
                        text=[f"{int(v)} Pcs" if v>0 else "" for v in dq.values],
                        textposition='outside', textfont=dict(size=9,color='#1a0030')))
                fig_szq_s.update_layout(**cl(360,f"Size-wise Qty Sold — {sz_filter_label}",margin=dict(l=10,r=10,t=55,b=40)),
                    barmode='group', bargap=0.15)
                st.plotly_chart(fig_szq_s, use_container_width=True)
        else:
            sz_sale_s = sdf_s.groupby('Size')['NetSale'].sum().reindex(ord_sz_s).fillna(0)
            sz_qty_s  = sdf_s.groupby('Size')['SaleQty'].sum().reindex(ord_sz_s).fillna(0)
            with sa_s:
                fig_szs_s = go.Figure(go.Bar(x=ord_sz_s, y=sz_sale_s.values,
                    marker=dict(color=sz_sale_s.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                    text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in sz_sale_s.values],
                    textposition='outside'))
                fig_szs_s.update_layout(**cl(320,f"Size-wise Net Sale — {sz_filter_label}",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                st.plotly_chart(fig_szs_s, use_container_width=True)
            with sb_s:
                fig_szq_s = go.Figure(go.Bar(x=ord_sz_s, y=sz_qty_s.values,
                    marker=dict(color=sz_qty_s.values, colorscale=[[0,'#fce7f3'],[0.5,'#ec4899'],[1,'#9d174d']], line=dict(width=0)),
                    text=[f"{int(v)} Pcs" if v>0 else "" for v in sz_qty_s.values],
                    textposition='outside'))
                fig_szq_s.update_layout(**cl(320,f"Size-wise Qty Sold — {sz_filter_label}",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                st.plotly_chart(fig_szq_s, use_container_width=True)

    with st6:
        sec("👤 Gender + Category Analysis")
        gf1_s,gf2_s,gf3_s,gf4_s = st.columns(4)
        with gf1_s: g_st_s = st.selectbox("Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="gc_st_s")
        with gf2_s: g_gen_s = st.selectbox("Gender", ["All"]+sorted(sale_s['Gender'].dropna().unique()), key="gc_gen_s")
        with gf3_s: g_cat_s = st.multiselect("Category", sorted(sale_s['Category'].dropna().unique()), key="gc_cat_s")
        with gf4_s:
            gcat_divs_s = sorted(sale_s[sale_s['Category'].isin(g_cat_s)]['DivisionDesc'].dropna().unique()) if g_cat_s else sorted(sale_s['DivisionDesc'].dropna().unique())
            g_div_s = st.multiselect("Division", gcat_divs_s, key="gc_div_s")
        gdf_s = sale_s.copy()
        if g_st_s != "All": gdf_s = gdf_s[gdf_s['StoreName']==g_st_s]
        if g_gen_s != "All": gdf_s = gdf_s[gdf_s['Gender']==g_gen_s]
        if g_cat_s: gdf_s = gdf_s[gdf_s['Category'].isin(g_cat_s)]
        if g_div_s: gdf_s = gdf_s[gdf_s['DivisionDesc'].isin(g_div_s)]
        ga_s,gb_s = st.columns(2)
        with ga_s:
            gdr_s2 = gdf_s.groupby('Gender')['NetSale'].sum().sort_values(ascending=False)
            gdr_s2 = gdr_s2[gdr_s2>0]
            fig_g_s = go.Figure(go.Pie(labels=gdr_s2.index.tolist(), values=gdr_s2.values.tolist(), hole=0.5,
                marker=dict(colors=['#7b1fa2','#e91e63','#1565c0'], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=12,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_g_s.update_layout(**cl(320,"Gender-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
            st.plotly_chart(fig_g_s, use_container_width=True)
        with gb_s:
            cat2_s = gdf_s.groupby('Category')['NetSale'].sum().sort_values(ascending=False)
            cat2_s = cat2_s[cat2_s>0]
            fig_c2_s = go.Figure(go.Bar(x=cat2_s.index.tolist(), y=cat2_s.values,
                marker=dict(color=CAT_COLORS[:len(cat2_s)], line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in cat2_s.values], textposition='outside'))
            fig_c2_s.update_layout(**cl(320,"Category-wise Sale",margin=dict(l=10,r=10,t=55,b=70)),
                bargap=0.3, xaxis_tickangle=-30)
            st.plotly_chart(fig_c2_s, use_container_width=True)
        sec("📦 Sub Category Analysis")
        subcat_s = gdf_s.groupby('SubCategory')['NetSale'].sum().sort_values(ascending=False)
        subcat_s = subcat_s[subcat_s>0]
        if len(subcat_s)>0:
            fig_sc_s = go.Figure(go.Bar(x=subcat_s.index.tolist(), y=subcat_s.values,
                marker=dict(color=CAT_COLORS[:len(subcat_s)], line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in subcat_s.values], textposition='outside'))
            fig_sc_s.update_layout(**cl(320,"Sub Category-wise Sale",margin=dict(l=10,r=10,t=55,b=90)),
                bargap=0.3, xaxis_tickangle=-35)
            st.plotly_chart(fig_sc_s, use_container_width=True)
        sec("📋 Gender × Category Table")
        gc_tbl_s = gdf_s.groupby(['Gender','Category','SubCategory']).agg(
            NetSale=('NetSale','sum'), Qty=('SaleQty','sum')
        ).reset_index().sort_values('NetSale',ascending=False)
        gc_tbl_s = gc_tbl_s[gc_tbl_s['NetSale']>0]
        gc_tbl_s['NetSale'] = gc_tbl_s['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
        gc_tbl_s.columns = ['Gender','Category','Sub Category','Net Sale','Qty']
        st.dataframe(gc_tbl_s, use_container_width=True, hide_index=True)

    with st7:
        sec("🗓️ Season-wise Sale Analysis")
        sea1_s,sea2_s = st.columns(2)
        with sea1_s: sea_st_s = st.selectbox("Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="sea_st_s")
        with sea2_s: sea_sel_s = st.multiselect("Season", sorted(sale_s['Season'].dropna().unique()),
                                                  default=sorted(sale_s['Season'].dropna().unique()), key="sea_sel_s")
        seadf_s = sale_s.copy()
        if sea_st_s != "All": seadf_s = seadf_s[seadf_s['StoreName']==sea_st_s]
        if sea_sel_s: seadf_s = seadf_s[seadf_s['Season'].isin(sea_sel_s)]
        sea_kpi_s = seadf_s.groupby('Season').agg(NetSale=('NetSale','sum'), Qty=('SaleQty','sum')).reset_index()
        sea_kpi_s = sea_kpi_s.sort_values('NetSale',ascending=False)
        cols_sea_s = st.columns(min(len(sea_kpi_s),5))
        for i,row in enumerate(sea_kpi_s.itertuples()):
            if i < len(cols_sea_s):
                with cols_sea_s[i]:
                    st.markdown(f'''<div style="background:linear-gradient(135deg,{CAT_COLORS[i%len(CAT_COLORS)]},#9c27b0);border-radius:14px;padding:1rem 1.2rem;margin-bottom:.5rem">
                        <div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.75)">🗓️ {row.Season}</div>
                        <div style="font-size:1.3rem;font-weight:800;color:#fff">₹{fmt_inr(int(row.NetSale))}</div>
                        <div style="font-size:.72rem;color:rgba(255,255,255,.7)">{int(row.Qty):,} Pcs</div>
                    </div>''', unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        sa7_s,sb7_s = st.columns(2)
        with sa7_s:
            sea_pie_s = seadf_s.groupby('Season')['NetSale'].sum().sort_values(ascending=False)
            fig_sp_s = go.Figure(go.Pie(labels=sea_pie_s.index.tolist(), values=sea_pie_s.values.tolist(), hole=0.5,
                marker=dict(colors=CAT_COLORS[:len(sea_pie_s)], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=12,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_sp_s.update_layout(**cl(320,"Season-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
            st.plotly_chart(fig_sp_s, use_container_width=True)
        with sb7_s:
            fig_sm_s = go.Figure()
            for i,sea in enumerate(sea_sel_s or []):
                sd7_s = seadf_s[seadf_s['Season']==sea].groupby('Month')['NetSale'].sum().reindex(MONTHS_S).fillna(0)
                fig_sm_s.add_trace(go.Scatter(x=MONTHS_S, y=sd7_s.values, name=sea,
                    mode='lines+markers', line=dict(color=CAT_COLORS[i%len(CAT_COLORS)],width=2.5), marker=dict(size=7)))
            fig_sm_s.update_layout(**cl(320,"Season Monthly Trend",margin=dict(l=10,r=10,t=55,b=40)))
            st.plotly_chart(fig_sm_s, use_container_width=True)

    with st8:
        sec("📊 Store: Sale vs Closing Stock")
        sv_st_s = st.selectbox("Filter Store", ["All"]+sorted(sale_s['StoreName'].dropna().unique()), key="svs_st_s")
        sv_sale_s  = sale_s.copy()
        sv_stk_s   = stock_s.copy()
        if sv_st_s != "All":
            sv_sale_s = sv_sale_s[sv_sale_s['StoreName']==sv_st_s]
            sv_stk_s  = sv_stk_s[sv_stk_s['StoreName']==sv_st_s]
        store_sale_t  = sv_sale_s.groupby('StoreName')['NetSale'].sum()
        store_stock_t = sv_stk_s.groupby('StoreName')['ClosingValue'].sum()
        stores_all_s  = sorted(set(store_sale_t.index) | set(store_stock_t.index))
        sv_vals_s = [float(store_sale_t.get(s,0)) for s in stores_all_s]
        sk_vals_s = [float(store_stock_t.get(s,0)) for s in stores_all_s]
        fig_svs_s = go.Figure()
        fig_svs_s.add_trace(go.Bar(name='Net Sale', x=[s[:25] for s in stores_all_s], y=sv_vals_s,
            marker_color='#7b1fa2', text=[f"₹{fmt_inr(int(v))}" for v in sv_vals_s],
            textposition='outside', textfont=dict(size=9,color='#1a0030')))
        fig_svs_s.add_trace(go.Bar(name='Closing Stock', x=[s[:25] for s in stores_all_s], y=sk_vals_s,
            marker_color='#ce93d8', text=[f"₹{fmt_inr(int(v))}" for v in sk_vals_s],
            textposition='outside', textfont=dict(size=9,color='#1a0030')))
        fig_svs_s.update_layout(**cl(420,"Store: Sale vs Closing Stock",margin=dict(l=10,r=10,t=55,b=110)),
            barmode='group', bargap=0.15, xaxis_tickangle=-45)
        st.plotly_chart(fig_svs_s, use_container_width=True)
        sec("📊 Sell-Through Rate by Store")
        str_data_s = []
        for s in stores_all_s:
            sv_v = float(store_sale_t.get(s,0))
            sk_v = float(store_stock_t.get(s,0))
            t_v  = sv_v + sk_v
            st_r = round(sv_v/t_v*100,1) if t_v>0 else 0
            str_data_s.append({'Store':s,'Sale':f"₹{fmt_inr(int(sv_v))}",'Stock':f"₹{fmt_inr(int(sk_v))}",
                'ST%':f"{st_r}%",'Status':'🟢 Good' if st_r>=60 else ('🟡 Avg' if st_r>=30 else '🔴 Low')})
        st.dataframe(pd.DataFrame(str_data_s), use_container_width=True, hide_index=True)

    with st9:
        sec("🏭 Warehouse Stock Analysis")
        if len(wh_s) > 0:
            wh_total_qty = int(wh_s['ClosingQty'].sum())
            wh_total_val = wh_s['ClosingValueMRP'].sum()
            wh1,wh2,wh3 = st.columns(3)
            kpi(wh1,"Total WH Qty",   f"{wh_total_qty:,} Pcs", "All godowns", "📦")
            kpi(wh2,"Total MRP Value", f"₹{fmt_inr(int(wh_total_val))}", "Closing stock", "💰")
            kpi(wh3,"Total Godowns",   str(wh_s['GodownName'].nunique()), "Active godowns", "🏭")
            st.markdown("<br>", unsafe_allow_html=True)
            sec("🏭 Godown-wise Stock")
            gd_s = wh_s.groupby('GodownName').agg(
                ClosingQty=('ClosingQty','sum'), ClosingValueMRP=('ClosingValueMRP','sum')
            ).reset_index().sort_values('ClosingValueMRP',ascending=False)
            wga,wgb = st.columns(2)
            with wga:
                fig_gd_s = go.Figure(go.Bar(
                    x=gd_s['ClosingValueMRP'].values,
                    y=[s[:25] for s in gd_s['GodownName'].tolist()], orientation='h',
                    marker=dict(color=gd_s['ClosingValueMRP'].values, colorscale=BLUE_SEQ, line=dict(width=0)),
                    text=[f"₹{fmt_inr(int(v))}" for v in gd_s['ClosingValueMRP'].values],
                    textposition='outside', textfont=dict(size=10,color='#1a0030')))
                fig_gd_s.update_layout(**cl(360,"Godown-wise Stock Value",margin=dict(l=10,r=160,t=55,b=40)),
                    xaxis_range=[0,gd_s['ClosingValueMRP'].max()*1.45])
                st.plotly_chart(fig_gd_s, use_container_width=True)
            with wgb:
                fig_gdq_s = go.Figure(go.Bar(
                    x=gd_s['ClosingQty'].values,
                    y=[s[:25] for s in gd_s['GodownName'].tolist()], orientation='h',
                    marker=dict(color=gd_s['ClosingQty'].values, colorscale=BLUE_SEQ, line=dict(width=0)),
                    text=[str(int(v)) for v in gd_s['ClosingQty'].values],
                    textposition='outside', textfont=dict(size=10,color='#1a0030')))
                fig_gdq_s.update_layout(**cl(360,"Godown-wise Stock Qty",margin=dict(l=10,r=100,t=55,b=40)),
                    xaxis_range=[0,gd_s['ClosingQty'].max()*1.35])
                st.plotly_chart(fig_gdq_s, use_container_width=True)
            sec("🗓️ Season-wise Warehouse Stock")
            wh_sea = wh_s.groupby('Season').agg(
                ClosingQty=('ClosingQty','sum'), ClosingValueMRP=('ClosingValueMRP','sum')
            ).reset_index().sort_values('ClosingValueMRP',ascending=False)
            whs_cols = st.columns(min(len(wh_sea),5))
            for i,row in enumerate(wh_sea.itertuples()):
                if i < len(whs_cols):
                    with whs_cols[i]:
                        st.markdown(f'''<div style="background:linear-gradient(135deg,{CAT_COLORS[i%len(CAT_COLORS)]},#9c27b0);border-radius:14px;padding:1rem 1.2rem;margin-bottom:.5rem">
                            <div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.75)">🗓️ {row.Season}</div>
                            <div style="font-size:1.1rem;font-weight:800;color:#fff">₹{fmt_inr(int(row.ClosingValueMRP))}</div>
                            <div style="font-size:.72rem;color:rgba(255,255,255,.7)">{int(row.ClosingQty):,} Pcs</div>
                        </div>''', unsafe_allow_html=True)
            st.markdown("<br>", unsafe_allow_html=True)
            sec("🏷️ Top 10 Articles by WH Stock")
            wh_arts = wh_s.groupby(['ItemID','ItemName'])['ClosingQty'].sum().sort_values(ascending=False).head(10)
            wh_arts_sorted = wh_arts.sort_values(ascending=True)
            if len(wh_arts_sorted)>0:
                fig_wart_s = go.Figure(go.Bar(
                    x=wh_arts_sorted.values,
                    y=[f"{idx[1][:22]}..{idx[0][-6:]}" for idx in wh_arts_sorted.index],
                    orientation='h',
                    marker=dict(color=wh_arts_sorted.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                    text=[str(int(v)) for v in wh_arts_sorted.values],
                    textposition='outside', textfont=dict(size=11,color='#1a0030')))
                fig_wart_s.update_layout(**cl(420,"Top 10 Articles by WH Stock Qty",margin=dict(l=10,r=100,t=55,b=40)),
                    xaxis_range=[0,wh_arts_sorted.max()*1.4])
                st.plotly_chart(fig_wart_s, use_container_width=True)
            FOOTWEAR_SIZES = ['6','7','8','9','10','11','12']
            CLOTHING_SIZES = ['XS','S','M','L','XL','XXL','XXXL','2XL','3XL','STANDARD']
            wh_foot  = wh_s[wh_s['Size'].isin(FOOTWEAR_SIZES)]
            wh_cloth = wh_s[wh_s['Size'].isin(CLOTHING_SIZES)]
            foot_sizes  = [s for s in FOOTWEAR_SIZES if s in wh_foot['Size'].unique()]
            cloth_sizes = [s for s in CLOTHING_SIZES if s in wh_cloth['Size'].unique()]
            if foot_sizes:
                sec("👟 Footwear Size-wise Warehouse Stock")
                foot_qty = wh_foot.groupby('Size')['ClosingQty'].sum().reindex(foot_sizes).fillna(0)
                foot_val = wh_foot.groupby('Size')['ClosingValueMRP'].sum().reindex(foot_sizes).fillna(0)
                wha_s,whb_s = st.columns(2)
                with wha_s:
                    fig_wsz_s = go.Figure(go.Bar(x=foot_sizes, y=foot_qty.values,
                        marker=dict(color=foot_qty.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                        text=[str(int(v)) if v>0 else "" for v in foot_qty.values], textposition='outside'))
                    fig_wsz_s.update_layout(**cl(320,"Footwear Size-wise WH Stock Qty",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                    st.plotly_chart(fig_wsz_s, use_container_width=True)
                with whb_s:
                    fig_wsv_s = go.Figure(go.Bar(x=foot_sizes, y=foot_val.values,
                        marker=dict(color=foot_val.values, colorscale=[[0,'#fce7f3'],[0.5,'#ec4899'],[1,'#9d174d']], line=dict(width=0)),
                        text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in foot_val.values], textposition='outside'))
                    fig_wsv_s.update_layout(**cl(320,"Footwear Size-wise WH Stock Value",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                    st.plotly_chart(fig_wsv_s, use_container_width=True)
            if cloth_sizes:
                sec("👕 Clothing Size-wise Warehouse Stock")
                cloth_qty = wh_cloth.groupby('Size')['ClosingQty'].sum().reindex(cloth_sizes).fillna(0)
                cloth_val = wh_cloth.groupby('Size')['ClosingValueMRP'].sum().reindex(cloth_sizes).fillna(0)
                whc_s,whd_s = st.columns(2)
                with whc_s:
                    fig_wcz_s = go.Figure(go.Bar(x=cloth_sizes, y=cloth_qty.values,
                        marker=dict(color=cloth_qty.values, colorscale=BLUE_SEQ, line=dict(width=0)),
                        text=[str(int(v)) if v>0 else "" for v in cloth_qty.values], textposition='outside'))
                    fig_wcz_s.update_layout(**cl(320,"Clothing Size-wise WH Stock Qty",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                    st.plotly_chart(fig_wcz_s, use_container_width=True)
                with whd_s:
                    fig_wcv_s = go.Figure(go.Bar(x=cloth_sizes, y=cloth_val.values,
                        marker=dict(color=cloth_val.values, colorscale=[[0,'#fce7f3'],[0.5,'#ec4899'],[1,'#9d174d']], line=dict(width=0)),
                        text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in cloth_val.values], textposition='outside'))
                    fig_wcv_s.update_layout(**cl(320,"Clothing Size-wise WH Stock Value",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                    st.plotly_chart(fig_wcv_s, use_container_width=True)
            sec("📋 Warehouse Stock Detail Table")
            wh_tbl_s = wh_s.groupby(['GodownName','ItemID','ItemName','Category','Gender','Size','Color','Season']).agg(
                ClosingQty=('ClosingQty','sum'), ClosingValueMRP=('ClosingValueMRP','sum')
            ).reset_index().sort_values('ClosingQty',ascending=False)
            wh_tbl_s['ClosingValueMRP'] = wh_tbl_s['ClosingValueMRP'].apply(lambda x: f"₹{fmt_inr(int(x))}")
            wh_tbl_s.columns = ['Godown','Item ID','Item Name','Category','Gender','Size','Colour','Season','Closing Qty','Closing Value']
            st.dataframe(wh_tbl_s, use_container_width=True, hide_index=True)
        else:
            st.info("No warehouse data found in Sheet2")

    with st10:
        sec("📥 Export Data")
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        def build_store_excel():
            wb = Workbook()
            def sc(ws, row, col, val, bg="FFFFFF", fg="1a0030", bold=False, sz=9):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(bold=bold, size=sz, color=fg.replace('#',''), name="Calibri")
                cell.fill = PatternFill("solid", fgColor=bg.replace('#',''))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin',color="e5e7eb"),right=Side(style='thin',color="e5e7eb"),
                    top=Side(style='thin',color="e5e7eb"),bottom=Side(style='thin',color="e5e7eb"))
            ws1_e = wb.active; ws1_e.title = "Store Summary"
            ws1_e.sheet_view.showGridLines = False
            hdrs = ['Store','Net Sale (₹)','Total Qty','MRP Value (₹)','Avg Disc%']
            for ci,h in enumerate(hdrs,1): sc(ws1_e,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
            st_data_e = sale_s.groupby('StoreName').agg(
                NetSale=('NetSale','sum'), Qty=('SaleQty','sum'),
                MRP=('MRPValue','sum'), Disc=('DiscAmt','sum')).reset_index().sort_values('NetSale',ascending=False)
            for ri,row in enumerate(st_data_e.itertuples(),2):
                dp = round(row.Disc/row.MRP*100,1) if row.MRP>0 else 0
                bg = "f5f3ff" if ri%2==0 else "FFFFFF"
                for ci,v in enumerate([row.StoreName,int(row.NetSale),int(row.Qty),int(row.MRP),f"{dp}%"],1):
                    sc(ws1_e,ri,ci,v,bg)
            for ci,w in enumerate([35,15,12,15,10],1):
                ws1_e.column_dimensions[get_column_letter(ci)].width = w
            ws2_e = wb.create_sheet("Monthly Store")
            ws2_e.sheet_view.showGridLines = False
            sc(ws2_e,1,1,"Store","1e3a5f","FFFFFF",bold=True,sz=10)
            for ci,m in enumerate(MONTHS_S,2): sc(ws2_e,1,ci,m,"1e3a5f","FFFFFF",bold=True,sz=9)
            sc(ws2_e,1,len(MONTHS_S)+2,"Total","1e3a5f","FFFFFF",bold=True,sz=10)
            piv_e = sale_s.pivot_table(index='StoreName',columns='Month',values='NetSale',aggfunc='sum').reindex(columns=MONTHS_S).fillna(0)
            piv_e['Total'] = piv_e.sum(axis=1)
            for ri,(sn,row) in enumerate(piv_e.iterrows(),2):
                sc(ws2_e,ri,1,sn,"f5f3ff" if ri%2==0 else "FFFFFF")
                for ci,v in enumerate(row.values,2):
                    sc(ws2_e,ri,ci,int(v) if v>0 else "","f5f3ff" if ri%2==0 else "FFFFFF")
            ws2_e.column_dimensions['A'].width = 35
            ws3_e = wb.create_sheet("Stock Summary")
            ws3_e.sheet_view.showGridLines = False
            for ci,h in enumerate(['Store','Closing Qty','Closing Value (₹)'],1): sc(ws3_e,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
            stk_e = stock_s.groupby('StoreName').agg(Qty=('ClosingQty','sum'),Val=('ClosingValue','sum')).reset_index().sort_values('Val',ascending=False)
            for ri,row in enumerate(stk_e.itertuples(),2):
                bg = "f5f3ff" if ri%2==0 else "FFFFFF"
                for ci,v in enumerate([row.StoreName,int(row.Qty),int(row.Val)],1): sc(ws3_e,ri,ci,v,bg)
            for ci,w in enumerate([35,12,18],1): ws3_e.column_dimensions[get_column_letter(ci)].width = w
            if len(wh_s)>0:
                ws4_e = wb.create_sheet("Warehouse Stock")
                ws4_e.sheet_view.showGridLines = False
                wh_hdrs = ['Godown','Item ID','Item Name','Category','Gender','Size','Season','Closing Qty','Closing Value MRP']
                for ci,h in enumerate(wh_hdrs,1): sc(ws4_e,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
                wh_e = wh_s.groupby(['GodownName','ItemID','ItemName','Category','Gender','Size','Season']).agg(
                    Qty=('ClosingQty','sum'),Val=('ClosingValueMRP','sum')).reset_index().sort_values('Qty',ascending=False)
                for ri,row in enumerate(wh_e.itertuples(),2):
                    bg = "f5f3ff" if ri%2==0 else "FFFFFF"
                    for ci,v in enumerate([row.GodownName,row.ItemID,row.ItemName,row.Category,row.Gender,row.Size,row.Season,int(row.Qty),int(row.Val)],1):
                        sc(ws4_e,ri,ci,v,bg)
                for ci,w in enumerate([22,20,30,15,10,10,10,12,18],1):
                    ws4_e.column_dimensions[get_column_letter(ci)].width = w
            out = BytesIO(); wb.save(out); out.seek(0)
            return out
        store_excel = build_store_excel()
        st.download_button("📥 Download Store Excel Report", data=store_excel,
            file_name="Dcyphr_Store_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        sec("📊 Raw Data Preview")
        prev_s = sale_s[['StoreName','Month','Category','SubCategory','Gender','Size','Color','Season','SaleQty','NetSale']].copy()
        prev_s.columns = ['Store','Month','Category','Sub Category','Gender','Size','Colour','Season','Sale Qty','Net Sale']
        st.dataframe(prev_s, use_container_width=True, hide_index=True)


    with st11:
        sec("🎯 Final Review — Store Business Summary")

        # 1. TOP & BOTTOM STORES
        st.markdown("### 🏆 Top & Bottom Performing Stores")
        store_perf = sale_s.groupby('StoreName')['NetSale'].sum().reset_index().sort_values('NetSale', ascending=False)
        store_perf['Sale'] = store_perf['NetSale'].apply(lambda x: fmt_inr(int(x)))
        top10_perf = store_perf.head(10).sort_values('NetSale', ascending=True)
        bot10_perf = store_perf.tail(10).sort_values('NetSale', ascending=False)

        col_top, col_bot = st.columns(2)
        with col_top:
            sec("🟢 Top 10 Stores")
            top_max = float(top10_perf['NetSale'].max()) if top10_perf['NetSale'].max()>0 else 1
            top_vals = [max(float(v), top_max*0.05) for v in top10_perf['NetSale'].values]
            fig_top = go.Figure(go.Bar(
                x=top_vals,
                y=[s[:25] for s in top10_perf['StoreName'].tolist()],
                orientation='h',
                marker=dict(color='#16a34a', line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in top10_perf['NetSale'].values],
                textposition='outside', textfont=dict(size=10,color='#1a0030')))
            fig_top.update_layout(**cl(420,"Top 10 Stores",margin=dict(l=10,r=160,t=40,b=20)),
                xaxis_range=[0, top_max*1.8])
            st.plotly_chart(fig_top, use_container_width=True)
        with col_bot:
            sec("🔴 Bottom 10 Stores")
            bot_abs_max = float(bot10_perf['NetSale'].abs().max()) if bot10_perf['NetSale'].abs().max()>0 else 1
            bot_vals = [max(float(abs(v)), bot_abs_max*0.08) for v in bot10_perf['NetSale'].values]
            bot_texts = [f"₹{fmt_inr(int(v))}" if abs(v)>=1 else "₹0" for v in bot10_perf['NetSale'].values]
            fig_bot = go.Figure(go.Bar(
                x=bot_vals,
                y=[s[:25] for s in bot10_perf['StoreName'].tolist()],
                orientation='h',
                marker=dict(color='#dc2626', line=dict(width=0)),
                text=bot_texts,
                textposition='outside', textfont=dict(size=10,color='#1a0030')))
            fig_bot.update_layout(**cl(420,"Bottom 10 Stores",margin=dict(l=10,r=160,t=40,b=20)),
                xaxis_range=[0, bot_abs_max*1.8])
            st.plotly_chart(fig_bot, use_container_width=True)

        st.markdown("---")

        # 2. TOP & BOTTOM STYLES
        st.markdown("### 👕 Top & Bottom Performing Styles")
        # Group by ItemName only (not ItemID) to avoid duplicate styles
        style_perf = sale_s.groupby('ItemName').agg(
            NetSale=('NetSale','sum'), Qty=('SaleQty','sum')
        ).reset_index().sort_values('NetSale', ascending=False)
        # Top 10 — highest on top
        top10_style = style_perf.head(10).sort_values('NetSale', ascending=True)
        # Bottom 10 — lowest on top
        bot10_style = style_perf.tail(10).sort_values('NetSale', ascending=False)

        cs1, cs2 = st.columns(2)
        with cs1:
            sec("🟢 Top 10 Styles")
            top_max = float(top10_style['NetSale'].max()) if top10_style['NetSale'].max()>0 else 1
            top_vals = [max(float(v), top_max*0.05) for v in top10_style['NetSale'].values]
            fig_ts = go.Figure(go.Bar(
                x=top_vals,
                y=[str(n)[:30] for n in top10_style['ItemName'].tolist()],
                orientation='h',
                marker=dict(color='#16a34a', line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" if v!=0 else "₹0" for v in top10_style['NetSale'].values],
                textposition='outside', textfont=dict(size=10,color='#1a0030')))
            fig_ts.update_layout(**cl(420,"Top 10 Styles",margin=dict(l=10,r=160,t=40,b=20)),
                xaxis_range=[0, top_max*1.8])
            st.plotly_chart(fig_ts, use_container_width=True)
        with cs2:
            sec("🔴 Bottom 10 Styles")
            bot_abs_max = float(bot10_style['NetSale'].abs().max()) if bot10_style['NetSale'].abs().max()>0 else 1
            bot_vals_s = [max(float(abs(v)), bot_abs_max*0.08) for v in bot10_style['NetSale'].values]
            bot_texts = []
            for v in bot10_style['NetSale'].values:
                if abs(v) < 1:
                    bot_texts.append("₹0")
                else:
                    bot_texts.append(f"₹{fmt_inr(int(v))}")
            fig_bs = go.Figure(go.Bar(
                x=bot_vals_s,
                y=[str(n)[:30] for n in bot10_style['ItemName'].tolist()],
                orientation='h',
                marker=dict(color='#dc2626', line=dict(width=0)),
                text=bot_texts,
                textposition='outside', textfont=dict(size=10,color='#1a0030')))
            fig_bs.update_layout(**cl(420,"Bottom 10 Styles",margin=dict(l=10,r=160,t=40,b=20)),
                xaxis_range=[0, bot_abs_max*1.8])
            st.plotly_chart(fig_bs, use_container_width=True)

        st.markdown("---")

        # 3. TOTAL SALE ACROSS MONTHS
        st.markdown("### 📅 Total Sale Across Months")
        monthly_fr = sale_s.groupby('Month')['NetSale'].sum().reindex(MONTHS_S).fillna(0)
        bi_fr = int(monthly_fr.values.argmax())
        wi_fr = int(monthly_fr.values.argmin())
        bcolors_fr = ['#9c27b0' if i not in [bi_fr,wi_fr] else ('#16a34a' if i==bi_fr else '#dc2626') for i in range(len(monthly_fr))]

        fig_mfr = go.Figure()
        fig_mfr.add_trace(go.Bar(
            x=MONTHS_S, y=monthly_fr.values, name='Net Sale',
            marker=dict(color=bcolors_fr, line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in monthly_fr.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_mfr.add_trace(go.Scatter(
            x=MONTHS_S, y=monthly_fr.values, name='Trend',
            mode='lines+markers', line=dict(color='#f59e0b', width=2.5, dash='dot'),
            marker=dict(size=7, color='#f59e0b')))
        fig_mfr.update_layout(**cl(340,"Monthly Net Sale — All Stores",margin=dict(l=10,r=10,t=55,b=40)),
            bargap=0.3, yaxis_range=[0, monthly_fr.max()*1.3])
        st.plotly_chart(fig_mfr, use_container_width=True)

        mk1,mk2,mk3,mk4 = st.columns(4)
        kpi(mk1,"Best Month", MONTHS_S[bi_fr], f"₹{fmt_inr(int(monthly_fr.values[bi_fr]))}", "🏆")
        kpi(mk2,"Lowest Month", MONTHS_S[wi_fr], f"₹{fmt_inr(int(monthly_fr.values[wi_fr]))}", "⬇️")
        avg_monthly = monthly_fr[monthly_fr>0].mean()
        kpi(mk3,"Avg Monthly Sale", f"₹{fmt_inr(int(avg_monthly))}", "Per month avg", "📊")
        mom_fr = monthly_fr.pct_change().dropna()
        last_mom = float(mom_fr.values[-1])*100 if len(mom_fr)>0 else 0
        kpi(mk4,"Last MoM Growth", f"{last_mom:+.1f}%", "vs previous month", "📈" if last_mom>=0 else "📉")

        st.markdown("---")

        # 4. HIGHLIGHTS & LOWLIGHTS
        st.markdown("### 💡 Highlights & Lowlights")
        hl1, hl2 = st.columns(2)

        with hl1:
            sec("✅ Highlights")
            highlights = []
            highlights.append(f"🏆 Best month: <b>{MONTHS_S[bi_fr]}</b> — ₹{fmt_inr(int(monthly_fr.values[bi_fr]))}")
            top_store_name = store_perf.iloc[0]['StoreName']
            top_store_sale = store_perf.iloc[0]['NetSale']
            highlights.append(f"🏪 Top store: <b>{top_store_name[:30]}</b> — ₹{fmt_inr(int(top_store_sale))}")
            if len(style_perf)>0:
                highlights.append(f"👕 Top style: <b>{style_perf.iloc[0]['ItemName'][:30]}</b> — ₹{fmt_inr(int(style_perf.iloc[0]['NetSale']))}")
            growth_months = [MONTHS_S[i+1] for i,v in enumerate(monthly_fr.pct_change().values[1:]) if v>0 and i+1<len(MONTHS_S)]
            if growth_months:
                highlights.append(f"📈 Growth months: <b>{', '.join(growth_months[:3])}</b>")
            top_cat_s = sale_s.groupby('Category')['NetSale'].sum()
            top_cat_s = top_cat_s[top_cat_s>0]
            if len(top_cat_s)>0:
                highlights.append(f"🗂️ Top category: <b>{top_cat_s.idxmax()}</b> — ₹{fmt_inr(int(top_cat_s.max()))}")
            for h in highlights:
                st.markdown(f'<div style="background:#f0fdf4;border-left:4px solid #16a34a;padding:.6rem 1rem;margin-bottom:.5rem;border-radius:0 8px 8px 0"><span style="font-size:.85rem;color:#1a0030">{h}</span></div>', unsafe_allow_html=True)

        with hl2:
            sec("⚠️ Lowlights")
            lowlights = []
            lowlights.append(f"⬇️ Weakest month: <b>{MONTHS_S[wi_fr]}</b> — ₹{fmt_inr(int(monthly_fr.values[wi_fr]))}")
            bot_store_name = store_perf.iloc[-1]['StoreName']
            bot_store_sale = store_perf.iloc[-1]['NetSale']
            lowlights.append(f"🏪 Lowest store: <b>{bot_store_name[:30]}</b> — ₹{fmt_inr(int(bot_store_sale))}")
            if len(style_perf)>0:
                lowlights.append(f"👕 Slowest style: <b>{style_perf.iloc[-1]['ItemName'][:30]}</b> — ₹{fmt_inr(int(style_perf.iloc[-1]['NetSale']))}")
            decline_months = [MONTHS_S[i+1] for i,v in enumerate(monthly_fr.pct_change().values[1:]) if v<0 and i+1<len(MONTHS_S)]
            if decline_months:
                lowlights.append(f"📉 Decline months: <b>{', '.join(decline_months[:3])}</b>")
            low_cont = store_perf[store_perf['NetSale'] < store_perf['NetSale'].sum()*0.01]
            if len(low_cont)>0:
                lowlights.append(f"⚠️ <b>{len(low_cont)} stores</b> contributing less than 1% of total sale")
            for l in lowlights:
                st.markdown(f'<div style="background:#fef2f2;border-left:4px solid #dc2626;padding:.6rem 1rem;margin-bottom:.5rem;border-radius:0 8px 8px 0"><span style="font-size:.85rem;color:#1a0030">{l}</span></div>', unsafe_allow_html=True)

        st.markdown("---")

        # 5. SELL THROUGH
        st.markdown("### 🔄 Sell Through Analysis")
        total_sale_fr = sale_s['NetSale'].sum()
        total_stock_fr = stock_s['ClosingValue'].sum()
        total_inv_fr = total_sale_fr + total_stock_fr
        overall_st = round(total_sale_fr/total_inv_fr*100, 1) if total_inv_fr>0 else 0

        st_k1, st_k2, st_k3 = st.columns(3)
        kpi(st_k1,"Overall Sell Through", f"{overall_st}%",
            "Good" if overall_st>=60 else ("Average" if overall_st>=30 else "Low"), "🔄")
        kpi(st_k2,"Total Sale Value", f"₹{fmt_inr(int(total_sale_fr))}", "Net sale", "💰")
        kpi(st_k3,"Closing Stock Value", f"₹{fmt_inr(int(total_stock_fr))}", "Unsold inventory", "📦")
        st.markdown("<br>", unsafe_allow_html=True)

        sec("🗓️ Season-wise Sell Through")
        sea_sale_fr = sale_s.groupby('Season')['NetSale'].sum()
        sea_stk_fr  = stock_s.groupby('Season')['ClosingValue'].sum() if 'Season' in stock_s.columns else pd.Series(dtype=float)
        # Only show seasons present in BOTH sale and stock
        common_seasons = sorted(set(sea_sale_fr.index) & set(sea_stk_fr.index))
        st_sea_data = []
        for sea in common_seasons:
            sv = float(sea_sale_fr.get(sea, 0))
            sk = float(sea_stk_fr.get(sea, 0))
            tv = sv + sk
            if tv <= 0 or sv < 0:
                str_pct = None  # N/A for negative or zero
            else:
                str_pct = round(sv/tv*100, 1)
            if str_pct is None:
                status = "N/A"
                color = '#6b7280'
                display = "N/A"
            elif str_pct >= 60:
                status = "Good"; color = '#16a34a'; display = f"{str_pct}%"
            elif str_pct >= 30:
                status = "Avg"; color = '#f59e0b'; display = f"{str_pct}%"
            else:
                status = "Low"; color = '#dc2626'; display = f"{str_pct}%"
            st_sea_data.append({'Season':sea,'display':display,'Sale':sv,'Stock':sk,'color':color,'Status':status})

        if st_sea_data:
            sea_cols_fr = st.columns(min(len(st_sea_data), 5))
            for i, row in enumerate(st_sea_data):
                if i < len(sea_cols_fr):
                    with sea_cols_fr[i]:
                        st.markdown(
                            '<div style="background:linear-gradient(135deg,' + row['color'] + ',#1a0030);border-radius:14px;padding:1rem 1.2rem;margin-bottom:.5rem">'
                            '<div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.75)">🗓️ ' + row['Season'] + '</div>'
                            '<div style="font-size:1.3rem;font-weight:800;color:#fff">' + row['display'] + '</div>'
                            '<div style="font-size:.72rem;color:rgba(255,255,255,.7)">Sale: ₹' + fmt_inr(int(row['Sale'])) + '</div>'
                            '<div style="font-size:.65rem;color:rgba(255,255,255,.6)">Stock: ₹' + fmt_inr(int(row['Stock'])) + ' · ' + row['Status'] + '</div>'
                            '</div>',
                            unsafe_allow_html=True)
        else:
            st.info("No matching seasons found between sale and stock data.")

        st.markdown("---")

        # 6. CUSTOMER PROFILING
        st.markdown("### 👤 Customer Profiling & Demographics")
        # Merge ACTIVEWEAR / ACTIVE WEAR duplicates
        sale_s_cp = sale_s.copy()
        sale_s_cp['Category'] = sale_s_cp['Category'].str.strip().str.replace(r'\s+', ' ', regex=True).str.upper()
        sale_s_cp['Category'] = sale_s_cp['Category'].replace({'ACTIVE WEAR': 'ACTIVEWEAR'})
        cp1, cp2 = st.columns(2)
        with cp1:
            sec("👤 Gender Split")
            gen_fr = sale_s_cp.groupby('Gender')['NetSale'].sum().sort_values(ascending=False)
            gen_fr = gen_fr[gen_fr>0]
            fig_gen_fr = go.Figure(go.Pie(
                labels=gen_fr.index.tolist(), values=gen_fr.values.tolist(), hole=0.55,
                marker=dict(colors=['#7b1fa2','#e91e63','#1565c0','#ff6f00'], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=12,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_gen_fr.update_layout(**cl(280,"Gender-wise Sale",margin=dict(l=10,r=10,t=40,b=10)))
            st.plotly_chart(fig_gen_fr, use_container_width=True)
        with cp2:
            sec("🗂️ Category Preference")
            cat_fr = sale_s_cp.groupby('Category')['NetSale'].sum().sort_values(ascending=False)
            cat_fr = cat_fr[cat_fr>0]
            fig_cat_fr = go.Figure(go.Pie(
                labels=cat_fr.index.tolist(), values=cat_fr.values.tolist(), hole=0.55,
                marker=dict(colors=CAT_COLORS[:len(cat_fr)], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=11,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_cat_fr.update_layout(**cl(280,"Category Preference",margin=dict(l=10,r=10,t=40,b=10)))
            st.plotly_chart(fig_cat_fr, use_container_width=True)

        sec("📐 Size Preference (Buyer Profile)")
        SIZE_ORDER_FR = ['XS','S','M','L','XL','XXL','XXXL','2XL','3XL','STANDARD']
        all_sz_fr = sale_s['Size'].dropna().unique().tolist()
        ord_sz_fr = [s for s in SIZE_ORDER_FR if s in all_sz_fr] + [s for s in all_sz_fr if s not in SIZE_ORDER_FR]
        sz_qty_fr = sale_s_cp.groupby('Size')['SaleQty'].sum().reindex(ord_sz_fr).fillna(0)
        fig_sz_fr = go.Figure(go.Bar(
            x=ord_sz_fr, y=sz_qty_fr.values,
            marker=dict(color=sz_qty_fr.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[str(int(v)) if v>0 else "" for v in sz_qty_fr.values],
            textposition='outside', textfont=dict(size=11,color='#1a0030')))
        fig_sz_fr.update_layout(**cl(280,"Size-wise Qty Sold",margin=dict(l=10,r=10,t=40,b=40)),
            bargap=0.3, yaxis_range=[0, max(sz_qty_fr.max()*1.25,1)])
        st.plotly_chart(fig_sz_fr, use_container_width=True)

        sec("🎨 Colour Preference")
        col_fr = sale_s_cp.groupby('Color')['SaleQty'].sum().sort_values(ascending=False).head(10)
        fig_col_fr = go.Figure(go.Bar(
            x=col_fr.index.tolist(), y=col_fr.values,
            marker=dict(color=CAT_COLORS[:len(col_fr)], line=dict(width=0)),
            text=[str(int(v)) for v in col_fr.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_col_fr.update_layout(**cl(280,"Top 10 Colours by Qty",margin=dict(l=10,r=10,t=40,b=70)),
            bargap=0.3, xaxis_tickangle=-30)
        st.plotly_chart(fig_col_fr, use_container_width=True)


    st.stop()

# ══════════════════════════════════════════
# CHANNEL MODE
# ══════════════════════════════════════════
df, MONTHS_ORDER, MONTH_SHORT = st.session_state.data
mode = st.session_state.mode

date_range = f"{MONTH_SHORT[0]}'{MONTHS_ORDER[0][-2:] if len(MONTHS_ORDER[0])>3 else '25'}–{MONTH_SHORT[-1]}'{MONTHS_ORDER[-1][-2:] if len(MONTHS_ORDER[-1])>3 else '26'}" if MONTHS_ORDER else "All Months"

total_sale = df['NetSale'].sum()
total_qty  = int(df['Sale Qty'].sum())
total_mrp  = df['MRP Value'].sum()
avg_disc   = (df['Discount Amount'].sum() / total_mrp * 100) if total_mrp > 0 else 0
top_ch     = df.groupby('Distributor')['NetSale'].sum().idxmax()
top_st     = df.groupby('Store Name')['NetSale'].sum().idxmax()

k1,k2,k3,k4,k5,k6 = st.columns(6)
kpi(k1, "Total Net Sale",   f"₹{fmt_inr(int(total_sale))}", date_range, "💰")
kpi(k2, "Total Qty Sold",   f"{total_qty:,} Pcs",           "All channels", "📦")
kpi(k3, "Avg Discount",     f"{avg_disc:.1f}%",             "On MRP value", "🏷️")
kpi(k4, "Total Channels",   str(df['Distributor'].nunique()), "Active distributors", "🔗")
kpi(k5, "Top Channel",      (top_ch[:20] if len(top_ch)>20 else top_ch), f"₹{fmt_inr(int(df.groupby('Distributor')['NetSale'].sum().max()))}", "🏆")
kpi(k6, "Top Store",        top_st[:18], f"₹{fmt_inr(int(df.groupby('Store Name')['NetSale'].sum().max()))}", "⭐")
st.markdown("<br>", unsafe_allow_html=True)

t1,t2,t3,t4,t5,t6,t7,t8,t9,t10 = st.tabs([
    "📈 Overview", "📦 Channel-wise", "🏪 Store-wise",
    "🗂️ Category & Gender", "🏷️ Article-wise",
    "🎨 Colour-wise", "📐 Size-wise",
    "👤 Gender+Category", "🗓️ Season-wise", "📋 Export"
])

with t1:
    sec("📈 Monthly Net Sale Trend")
    monthly = df.groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
    bi = int(monthly.values.argmax()); wi = int(monthly.values.argmin())
    bcolors = ['#9c27b0' if i not in [bi,wi] else ('#16a34a' if i==bi else '#dc2626') for i in range(len(monthly))]
    fig = go.Figure(go.Bar(
        x=MONTH_SHORT, y=monthly.values,
        marker=dict(color=bcolors, line=dict(width=0)),
        text=[f"₹{fmt_inr(int(v))}" for v in monthly.values],
        textposition='outside', textfont=dict(size=10,color='#1a0030'),
    ))
    fig.update_layout(**cl(360,f"Monthly Net Sale Trend — All Channels ({date_range})",
        margin=dict(l=10,r=10,t=55,b=40)),
        bargap=0.3, yaxis_range=[0, monthly.max()*1.25],
        annotations=[
            dict(x=MONTH_SHORT[bi],y=monthly.values[bi]*1.18,text="🏆 Best",showarrow=False,font=dict(color='#16a34a',size=10)),
            dict(x=MONTH_SHORT[wi],y=monthly.values[wi]*1.18,text="⬇ Low",showarrow=False,font=dict(color='#dc2626',size=10)),
        ])
    st.plotly_chart(fig, use_container_width=True)
    mom = monthly.pct_change()*100
    mp = [float(v) for v in mom.values[1:]]
    fig_mom = go.Figure(go.Bar(
        x=MONTH_SHORT[1:], y=mp,
        marker=dict(color=['#16a34a' if v>=0 else '#dc2626' for v in mp], line=dict(width=0)),
        text=[f"{v:+.1f}%" for v in mp], textposition='outside', textfont=dict(size=10,color='#1a0030')))
    fig_mom.update_layout(**cl(300,"Month-on-Month Growth (%)",margin=dict(l=10,r=10,t=55,b=40)), bargap=0.3)
    fig_mom.update_layout(yaxis=dict(gridcolor='#ede9fe',zeroline=True,zerolinecolor='#9c27b0',zerolinewidth=2))
    st.plotly_chart(fig_mom, use_container_width=True)
    ca,cb = st.columns(2)
    with ca:
        sec("🏆 Top Channels by Sale")
        ch_sale = df.groupby('Distributor')['NetSale'].sum().sort_values()
        fig2 = go.Figure(go.Bar(
            x=ch_sale.values, y=[c[:25] for c in ch_sale.index.tolist()], orientation='h',
            marker=dict(color=ch_sale.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in ch_sale.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig2.update_layout(**cl(320,"Channel-wise Total Net Sale",margin=dict(l=10,r=200,t=55,b=40)),
            xaxis_range=[0,ch_sale.max()*1.45])
        st.plotly_chart(fig2, use_container_width=True)
    with cb:
        sec("🗂️ Category Mix")
        cat_s = df.groupby('Category')['NetSale'].sum().sort_values(ascending=False)
        cat_s = cat_s[cat_s>0]
        fig3 = go.Figure(go.Pie(
            labels=cat_s.index.tolist(), values=cat_s.values.tolist(), hole=0.5,
            marker=dict(colors=CAT_COLORS[:len(cat_s)], line=dict(color='#fff',width=2)),
            textinfo='label+percent', textfont=dict(size=11,color='#1a0030'),
            insidetextfont=dict(size=10,color='#fff')))
        fig3.update_layout(**cl(320,"Category-wise Sale",margin=dict(l=10,r=10,t=55,b=10)),
            annotations=[dict(text=f"<b>₹{fmt_inr(int(total_sale))}</b>",x=0.5,y=0.5,
                font=dict(size=11,color='#1a0030'),showarrow=False)])
        st.plotly_chart(fig3, use_container_width=True)

with t2:
    sec("📦 Channel-wise Monthly Sale Trend")
    channels = df['Distributor'].dropna().unique().tolist()
    sel_ch = st.multiselect("Select Channels", channels, default=channels, key="ch_sel")
    if sel_ch:
        fig_ch = go.Figure()
        for i,ch in enumerate(sel_ch):
            cd = df[df['Distributor']==ch].groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
            fig_ch.add_trace(go.Bar(x=MONTH_SHORT, y=cd.values, name=ch[:25],
                marker_color=CAT_COLORS[i%len(CAT_COLORS)],
                hovertemplate=f'<b>{ch[:25]}</b><br>%{{x}}: ₹%{{y:,.0f}}<extra></extra>'))
        fig_ch.update_layout(**cl(400,"Channel-wise Monthly Sale",margin=dict(l=10,r=10,t=55,b=40)),
            barmode='group', bargap=0.12)
        st.plotly_chart(fig_ch, use_container_width=True)
    sec("📋 Channel Summary Table")
    ch_tbl = df.groupby('Distributor').agg(
        Total_Sale=('NetSale','sum'), Total_Qty=('Sale Qty','sum'),
        MRP_Value=('MRP Value','sum'), Discount=('Discount Amount','sum'),
        Stores=('Store Name','nunique')
    ).reset_index()
    ch_tbl['Avg Disc%']  = (ch_tbl['Discount']/ch_tbl['MRP_Value']*100).round(1)
    ch_tbl['Sale Cont%'] = (ch_tbl['Total_Sale']/ch_tbl['Total_Sale'].sum()*100).round(2)
    ch_tbl = ch_tbl.sort_values('Total_Sale',ascending=False)
    ch_tbl['Total_Sale'] = ch_tbl['Total_Sale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    ch_tbl['MRP_Value']  = ch_tbl['MRP_Value'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    ch_tbl['Total_Qty']  = ch_tbl['Total_Qty'].apply(int)
    ch_tbl.columns = ['Channel','Net Sale','Total Qty','MRP Value','Discount','Stores','Avg Disc%','Sale Cont%']
    ch_tbl = ch_tbl.drop(columns=['Discount'])
    st.dataframe(ch_tbl, use_container_width=True, hide_index=True)
    sec("📅 Channel Month-wise Detail")
    ch_pivot = df.pivot_table(index='Distributor',columns='Month',values='NetSale',aggfunc='sum').reindex(columns=MONTHS_ORDER).fillna(0)
    ch_pivot['Total'] = ch_pivot.sum(axis=1)
    ch_pivot = ch_pivot.sort_values('Total',ascending=False)
    ch_pivot = ch_pivot.map(lambda x: f"₹{fmt_inr(int(x))}" if x!=0 else "—")
    st.dataframe(ch_pivot, use_container_width=True)

with t3:
    sec("🔍 Store Deep Dive")
    stores = sorted(df['Store Name'].dropna().unique().tolist())
    sel_store = st.selectbox("Select Store", stores, key="st_dd")
    if sel_store:
        ss = df[df['Store Name']==sel_store]
        ts = ss['NetSale'].sum(); tq = int(ss['Sale Qty'].sum())
        rank = int(df.groupby('Store Name')['NetSale'].sum().rank(ascending=False)[sel_store])
        cont = ts/total_sale if total_sale>0 else 0
        m1,m2,m3,m4 = st.columns(4)
        kpi(m1,"Net Sale",    f"₹{fmt_inr(int(ts))}", date_range,"💰")
        kpi(m2,"Qty Sold",    f"{tq:,} Pcs","Total pieces","📦")
        kpi(m3,"Contribution",pct(cont,2),"Of total sale","📊")
        kpi(m4,"Store Rank",  f"#{rank}",f"Out of {df['Store Name'].nunique()} stores","🏅")
        st.markdown("<br>", unsafe_allow_html=True)
        da,db = st.columns([3,2])
        with da:
            mm = ss.groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
            fig_dm = go.Figure(go.Bar(x=MONTH_SHORT,y=mm.values,
                marker=dict(color=mm.values,colorscale=BLUE_SEQ,line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in mm.values],
                textposition='outside',textfont=dict(size=10,color='#1a0030')))
            fig_dm.update_layout(**cl(280,f"{sel_store} — Monthly Sale",margin=dict(l=10,r=10,t=50,b=40)),bargap=0.3)
            st.plotly_chart(fig_dm, use_container_width=True)
        with db:
            cd = ss.groupby('Category')['NetSale'].sum(); cd = cd[cd>0]
            if len(cd)>0:
                fig_dp = go.Figure(go.Pie(labels=cd.index.tolist(),values=cd.values.tolist(),hole=0.48,
                    marker=dict(colors=CAT_COLORS[:len(cd)],line=dict(color='#fff',width=2)),
                    textinfo='label+percent',textfont=dict(size=11,color='#1a0030'),
                    insidetextfont=dict(size=10,color='#fff')))
                fig_dp.update_layout(**cl(280,"Category Mix",margin=dict(l=10,r=10,t=50,b=10)))
                st.plotly_chart(fig_dp, use_container_width=True)
    st.markdown("---")
    sec("🏪 Top 10 Stores by Net Sale")
    top10_st = df.groupby('Store Name')['NetSale'].sum().nlargest(10).sort_values()
    fig_st = go.Figure(go.Bar(
        x=top10_st.values, y=[s[:30] for s in top10_st.index.tolist()], orientation='h',
        marker=dict(color=top10_st.values, colorscale=BLUE_SEQ, line=dict(width=0)),
        text=[f"₹{fmt_inr(int(v))}" for v in top10_st.values],
        textposition='outside', textfont=dict(size=10,color='#1a0030')))
    fig_st.update_layout(**cl(420,"Top 10 Stores by Net Sale",margin=dict(l=10,r=180,t=55,b=40)),
        xaxis_range=[0,top10_st.max()*1.45])
    st.plotly_chart(fig_st, use_container_width=True)
    sec("📋 Store Summary Table")
    st_tbl = df.groupby(['Distributor','Store Name']).agg(
        Total_Sale=('NetSale','sum'), Total_Qty=('Sale Qty','sum'),
        MRP_Value=('MRP Value','sum'), Discount=('Discount Amount','sum'),
    ).reset_index()
    st_tbl['Avg Disc%']  = (st_tbl['Discount']/st_tbl['MRP_Value']*100).round(1)
    st_tbl['Sale Cont%'] = (st_tbl['Total_Sale']/st_tbl['Total_Sale'].sum()*100).round(2)
    st_tbl = st_tbl.sort_values('Total_Sale',ascending=False)
    st_tbl['Total_Sale'] = st_tbl['Total_Sale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    st_tbl['MRP_Value']  = st_tbl['MRP_Value'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    st_tbl['Total_Qty']  = st_tbl['Total_Qty'].apply(int)
    st_tbl = st_tbl.drop(columns=['Discount'])
    st_tbl.columns = ['Channel','Store','Net Sale','Total Qty','MRP Value','Avg Disc%','Sale Cont%']
    st.dataframe(st_tbl, use_container_width=True, hide_index=True)

with t4:
    ca4,cb4 = st.columns(2)
    with ca4:
        sec("🗂️ Category Monthly Trend")
        cats = df['Category'].dropna().unique()
        fig_cat = go.Figure()
        for i,cat in enumerate(cats):
            cd = df[df['Category']==cat].groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
            fig_cat.add_trace(go.Scatter(x=MONTH_SHORT,y=cd.values,name=cat,
                mode='lines+markers',line=dict(color=CAT_COLORS[i%len(CAT_COLORS)],width=2.5),marker=dict(size=7)))
        fig_cat.update_layout(**cl(340,"Category Monthly Trend",margin=dict(l=10,r=10,t=55,b=40)))
        st.plotly_chart(fig_cat, use_container_width=True)
    with cb4:
        sec("👤 Gender Distribution")
        gdr = df.groupby('Gender')['NetSale'].sum().sort_values(ascending=False)
        gdr = gdr[gdr>0]
        fig_gdr = go.Figure(go.Pie(
            labels=gdr.index.tolist(),values=gdr.values.tolist(),hole=0.5,
            marker=dict(colors=['#7b1fa2','#e91e63','#1565c0'],line=dict(color='#fff',width=2)),
            textinfo='label+percent',textfont=dict(size=12,color='#1a0030'),
            insidetextfont=dict(size=10,color='#fff')))
        fig_gdr.update_layout(**cl(340,"Gender-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
        st.plotly_chart(fig_gdr, use_container_width=True)
    sec("🔍 Category × Gender × Channel Summary")
    cg_tbl = df.groupby(['Category','Gender','Distributor'])['NetSale'].sum().reset_index()
    cg_tbl = cg_tbl[cg_tbl['NetSale']>0].sort_values('NetSale',ascending=False)
    cg_tbl['NetSale'] = cg_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    cg_tbl.columns = ['Category','Gender','Channel','Net Sale']
    st.dataframe(cg_tbl, use_container_width=True, hide_index=True)
    sec("📅 Category Month-wise Table")
    cat_pivot = df.pivot_table(index='Category',columns='Month',values='NetSale',aggfunc='sum').reindex(columns=MONTHS_ORDER).fillna(0)
    cat_pivot['Total'] = cat_pivot.sum(axis=1)
    cat_pivot = cat_pivot.sort_values('Total',ascending=False)
    cat_pivot = cat_pivot.map(lambda x: f"₹{fmt_inr(int(x))}" if x!=0 else "—")
    st.dataframe(cat_pivot, use_container_width=True)

with t5:
    sec("🏷️ Article-wise Sale Analysis")
    af1,af2,af3 = st.columns(3)
    with af1: a_ch = st.selectbox("Channel", ["All"] + sorted(df['Distributor'].dropna().unique()), key="art_ch")
    with af2: a_cat = st.selectbox("Category", ["All"] + sorted(df['Category'].dropna().unique()), key="art_cat")
    with af3: a_search = st.text_input("🔍 Search Item Name", placeholder="e.g. Trackpants, Tee...", key="art_search")
    adf = df.copy()
    if a_ch != "All": adf = adf[adf['Distributor']==a_ch]
    if a_cat != "All": adf = adf[adf['Category']==a_cat]
    if a_search: adf = adf[adf['Item Name'].str.contains(a_search, case=False, na=False)]
    top_arts = adf.groupby(['Item ID','Item Name'])['NetSale'].sum().sort_values(ascending=False).head(10)
    sec("🏆 Top 10 Articles by Net Sale")
    if len(top_arts) > 0:
        top_arts_sorted = top_arts.sort_values(ascending=True)
        vals = list(top_arts_sorted.values)
        labs = [f"{idx[1][:22]}..{idx[0][-6:]}" for idx in top_arts_sorted.index]
        texts = [f"₹{fmt_inr(int(v))}" for v in vals]
        max_val = max(vals) if vals else 1
        fig_art = go.Figure(go.Bar(
            x=vals, y=labs, orientation='h',
            marker=dict(color=vals, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=texts, textposition='outside', textfont=dict(size=11, color='#1a0030')))
        fig_art.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030", family="Inter", size=12),
            margin=dict(l=10, r=160, t=55, b=20), height=520,
            title=dict(text="<b>Top 10 Articles by Net Sale</b>",
                font=dict(color="#1a0030", size=14, family="Plus Jakarta Sans")),
            xaxis=dict(range=[0, max_val*1.7], gridcolor="#ede9fe",
                tickfont=dict(color="#1a0030", size=11), showgrid=True),
            yaxis=dict(gridcolor="#ede9fe", tickfont=dict(color="#1a0030", size=11),
                showgrid=False, autorange='reversed'),
            showlegend=False)
        st.plotly_chart(fig_art, use_container_width=True)
    sec("📅 Article Monthly Trend")
    art_opts = [f"{idx[1]} ({idx[0]})" for idx in top_arts.index]
    if art_opts:
        sel_art = st.selectbox("Select Article", art_opts, key="sel_art")
        if sel_art:
            sel_item_id = sel_art.split('(')[-1].rstrip(')')
            art_monthly = adf[adf['Item ID']==sel_item_id].groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
            art_qty     = adf[adf['Item ID']==sel_item_id].groupby('Month')['Sale Qty'].sum().reindex(MONTHS_ORDER).fillna(0)
            am1,am2 = st.columns(2)
            with am1:
                fig_am = go.Figure(go.Bar(x=MONTH_SHORT, y=art_monthly.values,
                    marker=dict(color='#7b1fa2',line=dict(width=0)),
                    text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in art_monthly.values],
                    textposition='outside'))
                fig_am.update_layout(**cl(280,f"Monthly Sale — {sel_item_id}",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                st.plotly_chart(fig_am, use_container_width=True)
            with am2:
                fig_aq = go.Figure(go.Bar(x=MONTH_SHORT, y=art_qty.values,
                    marker=dict(color='#ce93d8',line=dict(width=0)),
                    text=[str(int(v)) if v>0 else "" for v in art_qty.values],
                    textposition='outside'))
                fig_aq.update_layout(**cl(280,f"Monthly Qty — {sel_item_id}",margin=dict(l=10,r=10,t=55,b=40)),bargap=0.3)
                st.plotly_chart(fig_aq, use_container_width=True)
    sec("📋 Article Summary Table")
    art_tbl = adf.groupby(['Item ID','Item Name','Category','Gender']).agg(
        NetSale=('NetSale','sum'), Qty=('Sale Qty','sum'),
        Channels=('Distributor','nunique'), Stores=('Store Name','nunique')
    ).reset_index().sort_values('NetSale',ascending=False)
    art_tbl['Sale Cont%'] = (art_tbl['NetSale']/art_tbl['NetSale'].sum()*100).round(2)
    art_tbl['NetSale'] = art_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    art_tbl['Qty'] = art_tbl['Qty'].apply(int)
    art_tbl.columns = ['Item ID','Item Name','Category','Gender','Net Sale','Qty','Channels','Stores','Sale Cont%']
    st.dataframe(art_tbl, use_container_width=True, hide_index=True)

with t6:
    sec("🎨 Item + Colour-wise Sale")
    cf1,cf2,cf3 = st.columns(3)
    with cf1: c_ch  = st.selectbox("Channel", ["All"]+sorted(df['Distributor'].dropna().unique()), key="col_ch")
    with cf2: c_cat = st.selectbox("Category", ["All"]+sorted(df['Category'].dropna().unique()), key="col_cat")
    with cf3: c_search = st.text_input("🔍 Search Item", placeholder="Item name...", key="col_search")
    cdf = df.copy()
    if c_ch != "All":  cdf = cdf[cdf['Distributor']==c_ch]
    if c_cat != "All": cdf = cdf[cdf['Category']==c_cat]
    if c_search: cdf = cdf[cdf['Item Name'].str.contains(c_search, case=False, na=False)]
    sec("🏆 Top Colours by Sale")
    col_sale = cdf.groupby('Color')['NetSale'].sum().sort_values(ascending=False).head(15)
    if len(col_sale)>0:
        fig_col = go.Figure(go.Bar(
            x=col_sale.index.tolist(), y=col_sale.values,
            marker=dict(color=CAT_COLORS[:len(col_sale)], line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in col_sale.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_col.update_layout(**cl(320,"Top Colours by Net Sale",margin=dict(l=10,r=10,t=55,b=70)),
            bargap=0.3, xaxis_tickangle=-30, yaxis_range=[0,col_sale.max()*1.25])
        st.plotly_chart(fig_col, use_container_width=True)
    sec("🔥 Item × Colour Sale Heatmap")
    top_items_c = cdf.groupby('Item Name')['NetSale'].sum().nlargest(10).index.tolist()
    top_cols_c  = cdf.groupby('Color')['NetSale'].sum().nlargest(10).index.tolist()
    hm_c = cdf[cdf['Item Name'].isin(top_items_c) & cdf['Color'].isin(top_cols_c)]
    hm_c = hm_c.pivot_table(index='Item Name',columns='Color',values='NetSale',aggfunc='sum').fillna(0)
    if len(hm_c)>0:
        fig_hmc = go.Figure(go.Heatmap(
            z=hm_c.values.tolist(), x=hm_c.columns.tolist(),
            y=[n[:30] for n in hm_c.index.tolist()],
            colorscale=[[0,'#fdf8ff'],[0.3,'#e9d8f8'],[0.7,'#9333ea'],[1,'#581c87']],
            text=[[f"₹{fmt_inr(int(v))}" if v>0 else "—" for v in row] for row in hm_c.values.tolist()],
            texttemplate="%{text}", textfont=dict(size=9,color='#1a0030'),
            colorbar=dict(title="Sale",tickfont=dict(color='#1a0030'))))
        fig_hmc.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030",family="Inter",size=10), height=420,
            margin=dict(l=10,r=20,t=55,b=80),
            title=dict(text="<b>Item × Colour Sale Heatmap (Top 10×10)</b>",
                       font=dict(color='#1a0030',size=13)),
            xaxis=dict(tickangle=-35,tickfont=dict(size=9)),
            yaxis=dict(tickfont=dict(size=9),autorange='reversed'))
        st.plotly_chart(fig_hmc, use_container_width=True)
    sec("📋 Item + Colour Detail Table")
    ic_tbl = cdf.groupby(['Item ID','Item Name','Color','Category']).agg(
        NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')
    ).reset_index().sort_values('NetSale',ascending=False)
    ic_tbl['NetSale'] = ic_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    ic_tbl['Qty'] = ic_tbl['Qty'].apply(int)
    ic_tbl.columns = ['Item ID','Item Name','Colour','Category','Net Sale','Qty']
    st.dataframe(ic_tbl, use_container_width=True, hide_index=True)

with t7:
    sec("📐 Size-wise Sale Analysis")
    sf1,sf2,sf3 = st.columns(3)
    with sf1: s_ch  = st.selectbox("Channel", ["All"]+sorted(df['Distributor'].dropna().unique()), key="sz_ch")
    with sf2: s_gen = st.selectbox("Gender", ["All"]+sorted(df['Gender'].dropna().unique()), key="sz_gen")
    with sf3: s_cat = st.selectbox("Category", ["All"]+sorted(df['Category'].dropna().unique()), key="sz_cat")
    sdf = df.copy()
    if s_ch  != "All": sdf = sdf[sdf['Distributor']==s_ch]
    if s_gen != "All": sdf = sdf[sdf['Gender']==s_gen]
    if s_cat != "All": sdf = sdf[sdf['Category']==s_cat]
    all_sizes = sdf['Size'].dropna().unique().tolist()
    ordered_sizes = [s for s in SIZE_ORDER if s in all_sizes] + [s for s in all_sizes if s not in SIZE_ORDER]
    sz_sale = sdf.groupby('Size')['NetSale'].sum().reindex(ordered_sizes).fillna(0)
    sz_qty  = sdf.groupby('Size')['Sale Qty'].sum().reindex(ordered_sizes).fillna(0)
    sa,sb = st.columns(2)
    with sa:
        sec("📊 Size-wise Net Sale")
        fig_szs = go.Figure(go.Bar(
            x=ordered_sizes, y=sz_sale.values,
            marker=dict(color=sz_sale.values, colorscale=BLUE_SEQ, line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" if v>0 else "" for v in sz_sale.values],
            textposition='outside', textfont=dict(size=11,color='#1a0030')))
        fig_szs.update_layout(**cl(320,"Size-wise Net Sale",margin=dict(l=10,r=10,t=55,b=40)),
            bargap=0.3, yaxis_range=[0,max(sz_sale.max()*1.25,1)])
        st.plotly_chart(fig_szs, use_container_width=True)
    with sb:
        sec("📦 Size-wise Qty Sold")
        fig_szq = go.Figure(go.Bar(
            x=ordered_sizes, y=sz_qty.values,
            marker=dict(color=sz_qty.values, colorscale=[[0,'#fce7f3'],[0.5,'#ec4899'],[1,'#9d174d']], line=dict(width=0)),
            text=[str(int(v)) if v>0 else "" for v in sz_qty.values],
            textposition='outside', textfont=dict(size=11,color='#1a0030')))
        fig_szq.update_layout(**cl(320,"Size-wise Qty Sold",margin=dict(l=10,r=10,t=55,b=40)),
            bargap=0.3, yaxis_range=[0,max(sz_qty.max()*1.25,1)])
        st.plotly_chart(fig_szq, use_container_width=True)
    sec("🔥 Size × Category Heatmap")
    sz_cat_hm = sdf.pivot_table(index='Category',columns='Size',values='Sale Qty',aggfunc='sum').reindex(columns=ordered_sizes).fillna(0)
    if len(sz_cat_hm)>0:
        fig_szhm = go.Figure(go.Heatmap(
            z=sz_cat_hm.values.tolist(), x=ordered_sizes, y=sz_cat_hm.index.tolist(),
            colorscale=[[0,'#fdf8ff'],[0.4,'#c084fc'],[1,'#581c87']],
            text=[[str(int(v)) if v>0 else "—" for v in row] for row in sz_cat_hm.values.tolist()],
            texttemplate="%{text}", textfont=dict(size=10,color='#1a0030'),
            colorbar=dict(title="Qty",tickfont=dict(color='#1a0030'))))
        fig_szhm.update_layout(
            paper_bgcolor="rgba(255,255,255,1)", plot_bgcolor="rgba(245,240,255,0.5)",
            font=dict(color="#1a0030",family="Inter",size=11), height=320,
            margin=dict(l=10,r=20,t=55,b=40),
            title=dict(text="<b>Size × Category Qty Sold</b>",font=dict(color='#1a0030',size=13)),
            xaxis=dict(tickfont=dict(size=10)), yaxis=dict(tickfont=dict(size=10),autorange='reversed'))
        st.plotly_chart(fig_szhm, use_container_width=True)
    sec("📅 Size Monthly Trend")
    fig_szm = go.Figure()
    for i,sz in enumerate(ordered_sizes[:8]):
        sd = sdf[sdf['Size']==sz].groupby('Month')['Sale Qty'].sum().reindex(MONTHS_ORDER).fillna(0)
        fig_szm.add_trace(go.Scatter(x=MONTH_SHORT, y=sd.values, name=sz,
            mode='lines+markers', line=dict(color=CAT_COLORS[i%len(CAT_COLORS)],width=2.5), marker=dict(size=7)))
    fig_szm.update_layout(**cl(320,"Size Monthly Qty Trend",margin=dict(l=10,r=10,t=55,b=40)))
    st.plotly_chart(fig_szm, use_container_width=True)
    sec("📋 Size Detail Table")
    sz_tbl = sdf.groupby(['Size','Category','Gender']).agg(
        NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')
    ).reset_index().sort_values('Qty',ascending=False)
    sz_tbl['Sale Cont%'] = (sz_tbl['NetSale']/sz_tbl['NetSale'].sum()*100).round(2)
    sz_tbl['NetSale'] = sz_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    sz_tbl['Qty'] = sz_tbl['Qty'].apply(int)
    sz_tbl.columns = ['Size','Category','Gender','Net Sale','Qty','Sale Cont%']
    st.dataframe(sz_tbl, use_container_width=True, hide_index=True)

with t8:
    sec("👤 Gender + Category + Sub Category Analysis")
    gf1,gf2 = st.columns(2)
    with gf1: g_ch  = st.selectbox("Channel", ["All"]+sorted(df['Distributor'].dropna().unique()), key="gc_ch")
    with gf2: g_gen = st.selectbox("Gender", ["All"]+sorted(df['Gender'].dropna().unique()), key="gc_gen")
    gdf = df.copy()
    if g_ch  != "All": gdf = gdf[gdf['Distributor']==g_ch]
    if g_gen != "All": gdf = gdf[gdf['Gender']==g_gen]
    ga,gb = st.columns(2)
    with ga:
        sec("👤 Gender Sale Distribution")
        gdr2 = gdf.groupby('Gender')['NetSale'].sum().sort_values(ascending=False)
        gdr2 = gdr2[gdr2>0]
        fig_g2 = go.Figure(go.Pie(
            labels=gdr2.index.tolist(), values=gdr2.values.tolist(), hole=0.5,
            marker=dict(colors=['#7b1fa2','#e91e63','#1565c0','#ff6f00'], line=dict(color='#fff',width=2)),
            textinfo='label+percent', textfont=dict(size=12,color='#1a0030'),
            insidetextfont=dict(size=10,color='#fff')))
        fig_g2.update_layout(**cl(320,"Gender-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
        st.plotly_chart(fig_g2, use_container_width=True)
    with gb:
        sec("🗂️ Category Sale Distribution")
        cat2 = gdf.groupby('Category')['NetSale'].sum().sort_values(ascending=False)
        cat2 = cat2[cat2>0]
        fig_c2 = go.Figure(go.Bar(
            x=cat2.index.tolist(), y=cat2.values,
            marker=dict(color=CAT_COLORS[:len(cat2)], line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in cat2.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_c2.update_layout(**cl(320,"Category-wise Sale",margin=dict(l=10,r=10,t=55,b=70)),
            bargap=0.3, xaxis_tickangle=-30, yaxis_range=[0,cat2.max()*1.25])
        st.plotly_chart(fig_c2, use_container_width=True)
    sec("📦 Sub Category Analysis")
    subcat = gdf.groupby('SubCategory')['NetSale'].sum().sort_values(ascending=False)
    subcat = subcat[subcat>0]
    if len(subcat)>0:
        fig_sc = go.Figure(go.Bar(
            x=subcat.index.tolist(), y=subcat.values,
            marker=dict(color=CAT_COLORS[:len(subcat)], line=dict(width=0)),
            text=[f"₹{fmt_inr(int(v))}" for v in subcat.values],
            textposition='outside', textfont=dict(size=10,color='#1a0030')))
        fig_sc.update_layout(**cl(320,"Sub Category-wise Sale",margin=dict(l=10,r=10,t=55,b=90)),
            bargap=0.3, xaxis_tickangle=-35, yaxis_range=[0,subcat.max()*1.25])
        st.plotly_chart(fig_sc, use_container_width=True)
    sec("✂️ Fit Type Analysis")
    fit = gdf.groupby('FitType')['NetSale'].sum().sort_values(ascending=False)
    fit = fit[fit>0]
    if len(fit)>0:
        ft1,ft2 = st.columns(2)
        with ft1:
            fig_fit = go.Figure(go.Pie(
                labels=fit.index.tolist(), values=fit.values.tolist(), hole=0.48,
                marker=dict(colors=CAT_COLORS[:len(fit)], line=dict(color='#fff',width=2)),
                textinfo='label+percent', textfont=dict(size=11,color='#1a0030'),
                insidetextfont=dict(size=10,color='#fff')))
            fig_fit.update_layout(**cl(300,"Fit Type Distribution",margin=dict(l=10,r=10,t=55,b=10)))
            st.plotly_chart(fig_fit, use_container_width=True)
        with ft2:
            fig_fitb = go.Figure(go.Bar(
                x=fit.index.tolist(), y=fit.values,
                marker=dict(color=CAT_COLORS[:len(fit)], line=dict(width=0)),
                text=[f"₹{fmt_inr(int(v))}" for v in fit.values],
                textposition='outside'))
            fig_fitb.update_layout(**cl(300,"Fit Type Sale",margin=dict(l=10,r=10,t=55,b=70)),
                bargap=0.3, xaxis_tickangle=-30)
            st.plotly_chart(fig_fitb, use_container_width=True)
    sec("📋 Gender × Category × Sub Category Table")
    gc_tbl = gdf.groupby(['Gender','Category','SubCategory']).agg(
        NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')
    ).reset_index().sort_values('NetSale',ascending=False)
    gc_tbl = gc_tbl[gc_tbl['NetSale']>0]
    gc_tbl['Sale Cont%'] = (gc_tbl['NetSale']/gc_tbl['NetSale'].sum()*100).round(2)
    gc_tbl['NetSale'] = gc_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    gc_tbl['Qty'] = gc_tbl['Qty'].apply(int)
    gc_tbl.columns = ['Gender','Category','Sub Category','Net Sale','Qty','Sale Cont%']
    st.dataframe(gc_tbl, use_container_width=True, hide_index=True)

with t9:
    sec("🗓️ Season-wise Sale Analysis")
    seasons = sorted(df['Season'].dropna().unique().tolist())
    seaf1,seaf2 = st.columns(2)
    with seaf1: sea_ch = st.selectbox("Channel", ["All"]+sorted(df['Distributor'].dropna().unique()), key="sea_ch")
    with seaf2: sea_sel = st.multiselect("Season", seasons, default=seasons, key="sea_sel")
    seadf = df.copy()
    if sea_ch != "All": seadf = seadf[seadf['Distributor']==sea_ch]
    if sea_sel: seadf = seadf[seadf['Season'].isin(sea_sel)]
    sea_kpi = seadf.groupby('Season').agg(NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')).reset_index()
    sea_kpi = sea_kpi.sort_values('NetSale',ascending=False)
    sea_cols = st.columns(len(sea_kpi))
    colors_sea = ['#7b1fa2','#1565c0','#2e7d32','#c62828']
    for i, row in sea_kpi.iterrows():
        col_idx = sea_kpi.index.tolist().index(i)
        if col_idx < len(sea_cols):
            with sea_cols[col_idx]:
                st.markdown(f"""<div style="background:linear-gradient(135deg,{colors_sea[col_idx%4]},#9c27b0);
                    border-radius:14px;padding:1rem 1.2rem;box-shadow:0 4px 16px rgba(0,0,0,.15);margin-bottom:.5rem">
                    <div style="font-size:.58rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:rgba(255,255,255,.75)">🗓️ {row['Season']}</div>
                    <div style="font-size:1.3rem;font-weight:800;color:#fff">₹{fmt_inr(int(row['NetSale']))}</div>
                    <div style="font-size:.72rem;color:rgba(255,255,255,.7)">{int(row['Qty']):,} Pcs</div>
                </div>""", unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    sa9,sb9 = st.columns(2)
    with sa9:
        sec("🗓️ Season Sale Distribution")
        sea_pie = seadf.groupby('Season')['NetSale'].sum().sort_values(ascending=False)
        fig_seap = go.Figure(go.Pie(
            labels=sea_pie.index.tolist(), values=sea_pie.values.tolist(), hole=0.5,
            marker=dict(colors=CAT_COLORS[:len(sea_pie)], line=dict(color='#fff',width=2)),
            textinfo='label+percent', textfont=dict(size=12,color='#1a0030'),
            insidetextfont=dict(size=10,color='#fff')))
        fig_seap.update_layout(**cl(320,"Season-wise Sale",margin=dict(l=10,r=10,t=55,b=10)))
        st.plotly_chart(fig_seap, use_container_width=True)
    with sb9:
        sec("📅 Season Monthly Trend")
        fig_seam = go.Figure()
        for i,sea in enumerate(sea_sel or seasons):
            sd9 = seadf[seadf['Season']==sea].groupby('Month')['NetSale'].sum().reindex(MONTHS_ORDER).fillna(0)
            fig_seam.add_trace(go.Scatter(x=MONTH_SHORT, y=sd9.values, name=sea,
                mode='lines+markers', line=dict(color=CAT_COLORS[i%len(CAT_COLORS)],width=2.5), marker=dict(size=7)))
        fig_seam.update_layout(**cl(320,"Season Monthly Sale Trend",margin=dict(l=10,r=10,t=55,b=40)))
        st.plotly_chart(fig_seam, use_container_width=True)
    sec("🗂️ Season × Category Breakdown")
    sea_cat = seadf.pivot_table(index='Season',columns='Category',values='NetSale',aggfunc='sum').fillna(0)
    fig_seacat = go.Figure()
    for i,cat in enumerate(sea_cat.columns):
        fig_seacat.add_trace(go.Bar(name=cat, x=sea_cat.index.tolist(), y=sea_cat[cat].values,
            marker_color=CAT_COLORS[i%len(CAT_COLORS)]))
    fig_seacat.update_layout(**cl(320,"Season × Category Sale",margin=dict(l=10,r=10,t=55,b=40)),
        barmode='group', bargap=0.2)
    st.plotly_chart(fig_seacat, use_container_width=True)
    sec("📦 Season × Channel Breakdown")
    sea_ch_tbl = seadf.pivot_table(index='Distributor',columns='Season',values='NetSale',aggfunc='sum').fillna(0)
    sea_ch_tbl['Total'] = sea_ch_tbl.sum(axis=1)
    sea_ch_tbl = sea_ch_tbl.sort_values('Total',ascending=False)
    sea_ch_tbl = sea_ch_tbl.map(lambda x: f"₹{fmt_inr(int(x))}" if x!=0 else "—")
    st.dataframe(sea_ch_tbl, use_container_width=True)
    sec("📋 Season Detail Table")
    sea_tbl = seadf.groupby(['Season','Category','Gender']).agg(
        NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')
    ).reset_index().sort_values('NetSale',ascending=False)
    sea_tbl = sea_tbl[sea_tbl['NetSale']>0]
    sea_tbl['Sale Cont%'] = (sea_tbl['NetSale']/sea_tbl['NetSale'].sum()*100).round(2)
    sea_tbl['NetSale'] = sea_tbl['NetSale'].apply(lambda x: f"₹{fmt_inr(int(x))}")
    sea_tbl['Qty'] = sea_tbl['Qty'].apply(int)
    sea_tbl.columns = ['Season','Category','Gender','Net Sale','Qty','Sale Cont%']
    st.dataframe(sea_tbl, use_container_width=True, hide_index=True)

with t10:
    sec("📥 Export Data")
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    def build_excel():
        wb = Workbook()
        def style_cell(ws, row, col, val, bg="FFFFFF", fg="1a0030", bold=False, sz=9):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(bold=bold, size=sz, color=fg.replace('#',''), name="Calibri")
            cell.fill = PatternFill("solid", fgColor=bg.replace('#',''))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin',color="e5e7eb"), right=Side(style='thin',color="e5e7eb"),
                top=Side(style='thin',color="e5e7eb"), bottom=Side(style='thin',color="e5e7eb"))
        ws1 = wb.active; ws1.title = "Channel Summary"
        ws1.sheet_view.showGridLines = False
        headers = ['Channel','Net Sale (₹)','Total Qty','MRP Value (₹)','Avg Disc%','Sale Cont%','Stores']
        for ci,h in enumerate(headers,1): style_cell(ws1,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
        ch_data = df.groupby('Distributor').agg(
            NetSale=('NetSale','sum'), Qty=('Sale Qty','sum'),
            MRP=('MRP Value','sum'), Disc=('Discount Amount','sum'),
            Stores=('Store Name','nunique')).reset_index().sort_values('NetSale',ascending=False)
        for ri,row in enumerate(ch_data.itertuples(),2):
            disc_pct = round(row.Disc/row.MRP*100,1) if row.MRP>0 else 0
            cont_pct = round(row.NetSale/total_sale*100,2) if total_sale>0 else 0
            vals = [row.Distributor, int(row.NetSale), int(row.Qty), int(row.MRP), f"{disc_pct}%", f"{cont_pct}%", row.Stores]
            bg = "f5f3ff" if ri%2==0 else "FFFFFF"
            for ci,v in enumerate(vals,1): style_cell(ws1,ri,ci,v,bg)
        for ci,w in enumerate([30,15,12,15,10,10,8],1):
            ws1.column_dimensions[get_column_letter(ci)].width = w
        ws2 = wb.create_sheet("Store Summary")
        ws2.sheet_view.showGridLines = False
        headers2 = ['Channel','Store','Net Sale (₹)','Total Qty','Avg Disc%','Sale Cont%']
        for ci,h in enumerate(headers2,1): style_cell(ws2,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
        st_data = df.groupby(['Distributor','Store Name']).agg(
            NetSale=('NetSale','sum'), Qty=('Sale Qty','sum'),
            MRP=('MRP Value','sum'), Disc=('Discount Amount','sum')).reset_index().sort_values('NetSale',ascending=False)
        for ri,row in enumerate(st_data.itertuples(),2):
            disc_pct = round(row.Disc/row.MRP*100,1) if row.MRP>0 else 0
            cont_pct = round(row.NetSale/total_sale*100,2) if total_sale>0 else 0
            store_name = getattr(row, 'Store_Name', getattr(row, '_3', ''))
            vals = [row.Distributor, store_name, int(row.NetSale), int(row.Qty), f"{disc_pct}%", f"{cont_pct}%"]
            bg = "f5f3ff" if ri%2==0 else "FFFFFF"
            for ci,v in enumerate(vals,1): style_cell(ws2,ri,ci,v,bg)
        for ci,w in enumerate([28,28,15,12,10,10],1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws3 = wb.create_sheet("Monthly Channel")
        ws3.sheet_view.showGridLines = False
        style_cell(ws3,1,1,"Channel","1e3a5f","FFFFFF",bold=True,sz=10)
        for ci,m in enumerate(MONTHS_ORDER,2): style_cell(ws3,1,ci,m,"1e3a5f","FFFFFF",bold=True,sz=9)
        style_cell(ws3,1,len(MONTHS_ORDER)+2,"Total","1e3a5f","FFFFFF",bold=True,sz=10)
        ch_pivot2 = df.pivot_table(index='Distributor',columns='Month',values='NetSale',aggfunc='sum').reindex(columns=MONTHS_ORDER).fillna(0)
        ch_pivot2['Total'] = ch_pivot2.sum(axis=1)
        for ri,(ch,row) in enumerate(ch_pivot2.iterrows(),2):
            style_cell(ws3,ri,1,ch,"f5f3ff" if ri%2==0 else "FFFFFF")
            for ci,v in enumerate(row.values,2):
                style_cell(ws3,ri,ci,int(v) if v>0 else "","f5f3ff" if ri%2==0 else "FFFFFF")
        ws3.column_dimensions['A'].width = 30
        for ci in range(2,len(MONTHS_ORDER)+3):
            ws3.column_dimensions[get_column_letter(ci)].width = 12
        ws4 = wb.create_sheet("Article Summary")
        ws4.sheet_view.showGridLines = False
        headers4 = ['Item ID','Item Name','Category','Gender','Net Sale (₹)','Qty','Sale Cont%']
        for ci,h in enumerate(headers4,1): style_cell(ws4,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
        art_data = df.groupby(['Item ID','Item Name','Category','Gender']).agg(
            NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')).reset_index().sort_values('NetSale',ascending=False)
        art_data['Cont'] = (art_data['NetSale']/art_data['NetSale'].sum()*100).round(2)
        for ri,row in enumerate(art_data.itertuples(),2):
            vals = [row._1, row._2, row.Category, row.Gender, int(row.NetSale), int(row.Qty), f"{row.Cont}%"]
            bg = "f5f3ff" if ri%2==0 else "FFFFFF"
            for ci,v in enumerate(vals,1): style_cell(ws4,ri,ci,v,bg)
        for ci,w in enumerate([25,35,15,10,15,10,10],1):
            ws4.column_dimensions[get_column_letter(ci)].width = w
        ws5 = wb.create_sheet("Season Summary")
        ws5.sheet_view.showGridLines = False
        headers5 = ['Season','Category','Gender','Net Sale (₹)','Qty','Sale Cont%']
        for ci,h in enumerate(headers5,1): style_cell(ws5,1,ci,h,"1e3a5f","FFFFFF",bold=True,sz=10)
        sea_data = df.groupby(['Season','Category','Gender']).agg(
            NetSale=('NetSale','sum'), Qty=('Sale Qty','sum')).reset_index().sort_values('NetSale',ascending=False)
        sea_data['Cont'] = (sea_data['NetSale']/sea_data['NetSale'].sum()*100).round(2)
        for ri,row in enumerate(sea_data.itertuples(),2):
            vals = [row.Season, row.Category, row.Gender, int(row.NetSale), int(row.Qty), f"{row.Cont}%"]
            bg = "f5f3ff" if ri%2==0 else "FFFFFF"
            for ci,v in enumerate(vals,1): style_cell(ws5,ri,ci,v,bg)
        for ci,w in enumerate([15,20,12,15,10,10],1):
            ws5.column_dimensions[get_column_letter(ci)].width = w
        out = BytesIO(); wb.save(out); out.seek(0)
        return out
    excel = build_excel()
    st.download_button(
        "📥  Download Excel Report",
        data=excel,
        file_name="Dcyphr_Sale_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False
    )
    st.markdown("<br>", unsafe_allow_html=True)
    sec("📊 Raw Data Preview")
    preview = df[['Distributor','Store Name','Month','Category','SubCategory','Gender','Size','Color','Season','Sale Qty','NetSale']].copy()
    preview['NetSale'] = preview['NetSale'].apply(lambda x: round(x,2))
    preview.columns = ['Channel','Store','Month','Category','Sub Category','Gender','Size','Colour','Season','Sale Qty','Net Sale']
    st.dataframe(preview, use_container_width=True, hide_index=True)
